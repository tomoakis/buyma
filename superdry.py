import unittest, time, requests, webbrowser, bs4, datetime, schedule, sys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import csv, os, re, xlrd, openpyxl
from selenium.webdriver.support.ui import WebDriverWait
import time, datetime, bs4, requests, urllib, xlwt, os, re, xlrd, openpyxl, unicodedata
from openpyxl.styles.fills import PatternFill

 

###############
### 基本設定 ###
###############

#エクセル指定
excelName = 'superdry'
xl_bk = xlrd.open_workbook(excelName + '.xlsx')
xl_sh = xl_bk.sheet_by_name(excelName)
wb = openpyxl.load_workbook('C:\\Users\\tomoa\\Desktop\\' + excelName + '.xlsx') 
ws = wb.active


#サイス取得開始番号の指定
browser = webdriver.Firefox()
browser.implicitly_wait(40)
time.sleep(1)
browser.get('https://www.google.com/')


def main():
    
    #列指定
    read_yoko = input('サイズ取得開始番号：')
    read_yoko = int(read_yoko) - 1

    yoko = input('新商品追加番号：')
    yoko = int(yoko) - 1
    folderNum = int(yoko) + 1

    print('')
    print('サイズ取得中')
    print('＝＝＝＝＝＝＝')
    
    while True:
##        try:
        print('')
        print(read_yoko + 1)

        #新商品の列に来たらbreak
        if read_yoko == yoko:
            break

        #URLへアクセス
        browser.implicitly_wait(40)
        time.sleep(1)
        ITEMURL = xl_sh.cell_value(read_yoko,6)
        browser.get(ITEMURL)
        browser.implicitly_wait(40)
        time.sleep(60)

        #URLが正しいか色で判断
        color = browser.find_element_by_xpath('//*[@id="template-container"]/section[1]/section[1]/section[3]/div/form/div[2]/h2[1]').text
        if color == '':
            color = browser.find_element_by_xpath('//*[@id="template-container"]/section[1]/section[1]/section[3]/div/form/div[2]/ul[1]/li/h2').text
        color = color.replace('Shown In ', '')
        print('既存カラー：' + xl_sh.cell_value(read_yoko,4))
        print('取得カラー：' + color)
        if not color == xl_sh.cell_value(read_yoko,4):
            ws['J' + str(int(read_yoko) + 1)].value = 'change'
            ws['I' + str(int(read_yoko) + 1)].value = 'change'
            print('change')
            
        else:
            sizebox = []
            sizeList = []
            size = browser.find_element_by_xpath('//*[@id="template-container"]/section[1]/section[1]/section[3]/div/form/div[2]/ul[2]').text
            sizebox.append(size)
            
            if sizebox == ['XS\nS\nM\nL\nXL\nXXL']:
                stop = 7
            elif sizebox == ['XXS\nXS\nS\nM\nL\nXL']:
                stop = 7
            elif sizebox == ['XS\nS\nM\nL\nXL']:
                stop = 6
            elif sizebox == ['XS\nS\nM\nL']:
                stop = 5
            elif sizebox == ['S\nM\nL\nXL']:
                stop = 5
            elif sizebox == ['ONE SIZE']:
                stop = 2
            elif sizebox == ['26W\n28W\n29W\n30W\n31W\n32W\n33W\n34W\n36W\n38W']:
                stop = 11
            elif sizebox == ['26W\n28W\n29W\n30W\n31W\n32W\n33W\n34W\n36W']:
                stop = 10
            elif sizebox == ['28W\n30W\n31W\n32W\n33W\n34W\n36W\n38W']:
                stop = 9
            elif sizebox == ['28W\n30W\n32W\n34W']:
                stop = 5
            elif sizebox == ['00\n0\n1\n3\n5\n7\n9\n11\n13\n15']:
                stop = 11
            elif sizebox == ['5/6\n7/8\n9/10\n11/12\n13/14\n15/16']:
                stop = 7
                
            num = 1
            while True:
                if num == stop:
                    break
                size = browser.find_element_by_xpath('//*[@id="template-container"]/section[1]/section[1]/section[3]/div/form/div[2]/ul[2]/li[' + str(num) + ']')
                sizes = size.get_attribute('class')
                if sizes == 'product-attrs__attr':
                    SIZE = browser.find_element_by_xpath('//*[@id="template-container"]/section[1]/section[1]/section[3]/div/form/div[2]/ul[2]/li[' + str(num) + ']').text
                    sizeList.append(SIZE)
                elif sizes == 'product-attrs__attr selected product-attrs__large-attr':
                    SIZE = browser.find_element_by_xpath('//*[@id="template-container"]/section[1]/section[1]/section[3]/div/form/div[2]/ul[2]/li[' + str(num) + ']').text
                    sizeList.append(SIZE)
                elif sizes == 'product-attrs__attr  product-attrs__large-attr':
                    SIZE = browser.find_element_by_xpath('//*[@id="template-container"]/section[1]/section[1]/section[3]/div/form/div[2]/ul[2]/li[' + str(num) + ']').text
                    sizeList.append(SIZE)

                num += 1
                    
            sizeList = ','.join(sizeList)
            ws['I' + str(int(read_yoko) + 1)].value = sizeList
            if sizeList == xl_sh.cell_value(read_yoko,8):
                ws['J' + str(int(read_yoko) + 1)].value = '-'
            if not sizeList == xl_sh.cell_value(read_yoko,8):
                ws['I' + str(int(read_yoko) + 1)].value = sizeList
                ws['J' + str(int(read_yoko) + 1)].value = 'yes'
            print(sizeList)
            
        read_yoko += 1
        wb.save('C:\\Users\\tomoa\\Desktop\\' + excelName + '.xlsx')

##        except:
##            ws['J' + str(int(read_yoko) + 1)].value = 'allNo'
##            ws['I' + str(int(read_yoko) + 1)].value = 'allNo'
##            print('allNo')
##            wb.save('C:\\Users\\tomoa\\Desktop\\' + excelName + '.xlsx')
##            read_yoko += 1




###########################################################################################################
###########################################################################################################


    #新商品追加
    print('')
    print('新商品追加')
    print('＝＝＝＝＝＝＝')
    print('')
    while True:
        print('')
        print(yoko + 1)
##        try:
        #URLへ移動
        browser.implicitly_wait(40)
        time.sleep(1)
        ITEMURL = xl_sh.cell_value(yoko,4)
        if ITEMURL == '-':
            break
        browser.get(ITEMURL)
        browser.implicitly_wait(40)
        time.sleep(10)


        #商品名
        name = browser.find_element_by_xpath('/html/body/div[4]/div/div[1]/div/div/div/div[2]/div/h1').text
        ws['A' + str(int(yoko) + 1)].value = name
        japanese(name, yoko, ws)
        ws['B' + str(int(yoko) + 1)].value = 'Superdry★' + japanese(name, yoko, ws)
        print(name)

        #値段
        price = browser.find_element_by_xpath('/html/body/div[4]/div/div[1]/div/div/div/div[2]/div/div[1]').text
        price = price.replace('USD $', '')
        ws['D' + str(int(yoko) + 1)].value = price
        print(price)

                
##        #カラー選択
##        colornum = 1
##        while True:
##            try:
##                browser.find_element_by_xpath('/html/body/div[4]/div/div[1]/div/div/div/div[2]/div/div[7]/div/div[2]/div/div[' + str(colornum) + ']/a/img').click()
##                time.sleep(5)
##            except:
##                break
##                              
##
##
##            #色
##            if colornum == 1:
##                colorcolum = 'H'
##                transcolor = 'I'
##                sizecolum = 'J'
##            elif colornum == 2:
##                colorcolum = 'K'
##                transcolor = 'L'
##                sizecolum = 'M'
##            elif colornum == 3:
##                colorcolum = 'N'
##                transcolor = 'O'
##                sizecolum = 'P'
##            elif colornum == 4:
##                colorcolum = 'Q'
##                transcolor = 'R'
##                sizecolum = 'S'
##            elif colornum == 5:
##                colorcolum = 'T'
##                transcolor = 'U'
##                sizecolum = 'V'
##            elif colornum == 6:
##                colorcolum = 'W'
##                transcolor = 'X'
##                sizecolum = 'Y'
##            elif colornum == 7:
##                colorcolum = 'Z'
##                transcolor = 'AA'
##                sizecolum = 'AB'
##            elif colornum == 8:
##                colorcolum = 'AC'
##                transcolor = 'AD'
##                sizecolum = 'AE'
##            elif colornum == 9:
##                colorcolum = 'AF'
##                transcolor = 'AG'
##                sizecolum = 'AH'
##            elif colornum == 10:
##                colorcolum = 'AI'
##                transcolor = 'AJ'
##                sizecolum = 'AK'

        color = browser.find_element_by_xpath('/html/body/div[4]/div/div[1]/div/div/div/div[2]/div/div[6]/div').text
        color = color.replace('Color: ', '')
        ws[colorcolum + str(int(yoko) + 1)].value = color
        colorsTranslate(color)
        ws[transcolor + str(int(yoko) + 1)].value = colorsTranslate(color)
        print(color)

            
##            #サイズ
##            sizeList = []
##            num = 1
##            try:
##                while True:
##                    size = browser.find_element_by_xpath('/html/body/div[2]/div/div/div/div/div[1]/div[2]/div/div[4]/div[1]/div[2]/div[' + str(num) + ']')
##                    sizes = size.get_attribute('class')
##                    if 'size-box' in sizes and not 'size-box invalid' in sizes:
##                        SIZE = browser.find_element_by_xpath('//*[@id="add_to_bag_desktop"]/div[1]/div[2]/div[' + str(num) + ']').text
##                        sizeList.append(SIZE)
##
##                    num += 1
##                    
##            except:  
##                sizeList = ','.join(sizeList)
##                ws[sizecolum + str(int(yoko) + 1)].value = sizeList
##                print(sizeList)


        
        #画像
        num = 2
        try:
            while True:
                if num == 5:
                    break
                img = browser.find_element_by_xpath('/html/body/div[4]/div/div[1]/div/div/div/div[1]/div[2]/div/div[2]/div/div[2]/div/div[' + str(num) + ']/img')
                src = img.get_attribute('src')
                urllib.request.urlretrieve(src, 'C:\\Users\\tomoa\\Desktop\\' + excelName + '\\' + str(folderNum) + '\\' + colorcolum + str(num) + ".jpg")
                num += 1
            print('画像取得完了')
        except:
            print('画像取得完了')

        
##            colornum  += 1



        ws['F' + str(int(yoko) + 1)].value = 'new'
        ws['AM' + str(int(yoko) + 1)].value = 'Superdry(極度乾燥しなさい)'
        wb.save('C:\\Users\\tomoa\\Desktop\\' + excelName + '.xlsx')
        yoko += 1
        folderNum = yoko + 1
                
                        

                    
##        except IndexError:
##            time.sleep(5)
##            browser.close()
##            time.sleep(5)
##            sys.exit()
##        
##        except:
##            ws['F' + str(int(yoko) + 1)].value = 'エラー'
##            ws['AM' + str(int(yoko) + 1)].value = 'Superdry(極度乾燥しなさい)'
##            wb.save('C:\\Users\\tomoa\\Desktop\\' + excelName + '.xlsx')
##            yoko += 1
##            folderNum = yoko + 1
##            print('エラー')
##            print('-----------------------')







########################################################################################################################
########################################################################################################################


# 色の翻訳
def colorsTranslate(color):
    
    if "lack" in color:
        return("ブラック")
    elif "&" in color:
        return("マルチ")
    elif "White" in color:
        return("ホワイト")
    elif "Cream" in color:
        return("ホワイト")
    elif "Brown" in color:
        return("ブラウン")
    elif "Chocolate" in color or 'Chestnut' in color or 'Mustard' in color:
        return("ブラウン")
    elif "Ivory" in color:
        return("グレー")
    elif "Tan" in color:
        return("グレー")
    elif "Navy" in color:
        return("ネイビー")
    elif "Teal" in color:
        return("ネイビー")
    elif "Gold" in color:
        return("イエロー")
    elif "Yellow" in color:
        return("イエロー")
    elif "Green" in color:
        return("グリーン")
    elif "Lime" in color:
        return("グリーン")
    elif "Olive" in color:
        return("グリーン")
    elif "Pink" in color:
        return("ピンク")
    elif "Peach" in color:
        return("ピンク")
    elif "lue" in color:
        return("ブルー")
    elif "Turquoise" in color:
        return("グリーン")
    elif "Denim" in color:
        return("ブルー")
    elif "Sky" in color or 'Slate' in color:
        return("ブルー")
    elif "Grey" in color or 'Lilac' in color:
        return("グレー")
    elif "Charcoal" in color:
        return("グレー")
    elif "Taupe" in color:
        return("グレー")
    elif "Red" in color or 'Rust' in color or 'Crimson' in color:
        return("レッド")
    elif "Maroon" in color:
        return("レッド")
    elif "Coral" in color:
        return("レッド")
    elif "Rose" in color:
        return("レッド")
    elif "Orange" in color:
        return("オレンジ")
    elif "Purple" in color:
        return("パープル")
    elif "Violet" in color:
        return("パープル")
    elif "Lavender" in color:
        return("パープル")
    elif "Beige" in color or 'Khaki' in color:
        return("ベージュ")
    elif "Indigo" in color:
        return("ブルー")
    elif "Assorted" in color:
        return("マルチカラー")
    elif "Burgundy" in color:
        return("レッド")
    elif "Camo" in color:
        return("マルチカラー")
    elif "Mauve" in color:
        return("マルチカラー")
    elif "Multi" in color:
        return("マルチカラー")
    elif "Berry" in color:
        return("ピンク")
    elif "Neutral" in color or 'Birch' in color:
        return("ホワイト")
    elif "Silver" in color or 'Mint' in color or 'Platinum' in color:
        return("シルバー")
    elif "Honey" in color or 'Mint' in color:
        return("マルチカラー")
    else:
        return("#################")


def japanese(name, NewProductNum, ws):
    title = []

    split = name.split()
    for i in split:

        if i == "Slide":
            title.append('シャワー')
        elif i == 'Denim':
            title.append('デニム')
        elif i == "Flip-Flop":
            title.append('ビーチサンダル')
        elif i == 'Cardigan':
            title.append('カーディガン')
        elif i == "Sweater":
            title.append('セーター')
        elif i == "Hoodie":
            title.append('パーカー')
        elif i == "Jacket" or i == 'Coat' or i == 'Vest' or i == 'Overshirt':
            title.append('ジャケット')
        elif i == "Tee":
            title.append('Tシャツ')
        elif i == "Jersey":
            title.append('ジャージ')
        elif i == "Shirt":
            title.append('シャツ')
        elif i == "Sock":
            title.append('ソックス/靴下')
        elif i == "Boxer":
            title.append('ボクサーパンツ')
        elif i == "Brief":
            title.append('ブリーフ')
        elif i == "Sweatshirt":
            title.append('スウェットシャツ')
        elif i == "Jogger":
            title.append('ジョガー')
        elif i == "Pants":
            title.append('パンツ')
        elif i == "Skinny":
            title.append('スキニー')
        elif i == "Ankle":
            title.append('アンクル')
        elif i == "Shorts":
            title.append('ショートパンツ')
        elif i == "Hat":
            title.append('キャップ')
        elif i == "Flannel":
            title.append('フランネル')
        elif i == "Print":
            title.append('プリント')
        elif i == "Waffle":
            title.append('ワッフル')
        elif i == "Beanie":
            title.append('ニット帽')
        elif i == "Sweatpants":
            title.append('スウェットパンツ')
        elif i == "Gazelle":
            title.append('ガゼル')
        elif i == "Sneaker":
            title.append('スニーカー')
        elif i == "Superstar":
            title.append('スーパースター')
        elif i == "Sandals":
            title.append('サンダル')
        elif i == "Fur":
            title.append('ファー')
        elif i == "Leather":
            title.append('レザー')
        elif i == "Vintage":
            title.append('ビンテージ')
        elif i == "Suede":
            title.append('スエード')
        elif i == "Pastel":
            title.append('パステル')
        elif i == "Mule":
            title.append('ミュール')
        elif i == "Heel":
            title.append('ハイヒール')
        elif i == "Boot":
            title.append('ブーツ')
        elif i == "Sunglasses":
            title.append('サングラス')
        elif i == "Bag":
            title.append('バッグ')
        elif i == "Necklace" or i == "Bracelet":
            title.append('ネックレス')
        elif i == "Backpack":
            title.append('バックパック')
        elif i == "Robe":
            title.append('ローブ')
        elif i == "Watch":
            title.append('時計')
        elif i == "Ring":
            title.append('リング') 
        elif i == "Tank":
            title.append('タンク')
        elif i == "Top":
            title.append('トップ')
        elif i == "Overall":
            title.append('オーバーオール')
        elif i == "Leggings":
            title.append('レギンス')
        elif i == "Track":
            title.append('トラック')
        elif i == "High-Rise":
            title.append('ハイライズ')
        elif i == "Active":
            title.append('アクティブ')
        elif i == "Mini":
            title.append('ミニ')
        elif i == "Twil":
            title.append('ツイル')
        elif i == "Bralette":
            title.append('ブラ')
        elif i == "Hipster":
            title.append('ヒップスターショーツ')
        elif i == "Bra":
            title.append('ブラ')
        elif i == "Undie":
            title.append('下着パンツ')
        elif i == "Cotton":
            title.append('コットン')
        elif i == "Thong":
            title.append('下着パンツ')
        elif i == "Tanga":
            title.append('下着パンツ')
        elif i == "Body":
            title.append('ボディ')
        elif i == "Bikini":
            title.append('ビキニ')
        elif i == "Top":
            title.append('トップ')
        elif i == "Bottom":
            title.append('ボトムパンツ')
        elif i == "One-Piece":
            title.append('ワンピース')
        elif i == "Swimsuit":
            title.append('水着')
        elif i == "Duffle":
            title.append('ダッフル')
        elif i == "Stripe":
            title.append('ストライプ')
        elif i == "Belt":
            title.append('ベルト')
        elif i == "Crossbody":
            title.append('ショルダー')
        elif i == "Tote":
            title.append('トート')
        elif i == "Canvas":
            title.append('キャンバス')
        elif i == "Wedge":
            title.append('ウェッジ')
        elif i == "Button-Down":
            title.append('ボタン')
        elif i == "Satin":
            title.append('サテン')
        elif i == "Fleece":
            title.append('フリース')
        elif i == "Logo":
            title.append('ロゴ')
        elif i == "Pocket":
            title.append('ポケット')
        elif i == "Bomber":
            title.append('ボンバー')
        elif i == "Hooded":
            title.append('フード付き')
        elif i == "Parka":
            title.append('パーカー')
        elif i == "Coach":
            title.append('コーチ')
        elif i == "Windbreaker":
            title.append('ウィンドブレーカー')
        elif i == "Popover":
            title.append('ポップオーバー')
        elif i == "Polo":
            title.append('ポロ')
        elif i == "Boxer":
            title.append('ボクサー')
        elif i == "Brief":
            title.append('ブリーフ')
        elif i == "Floral":
            title.append('花柄')
        elif i == "Full-Zip":
            title.append('ジップアップ')
        elif i == "Set":
            title.append('セット')
        elif i == "Sleep":
            title.append('パジャマ')
        elif i == "Flannele":
            title.append('フランネル')
        elif i == "Plaid":
            title.append('チェック')
        elif i == "Pocket":
            title.append('ポケット')
        elif i == "Checkerboard":
            title.append('チェック')
        elif i == "Patterned":
            title.append('柄入り')
        elif i == "Oxford ":
            title.append('オックスフォード')
        elif i == "Poplin ":
            title.append('アイコン')
        elif i == "Faux":
            title.append('フォックス')
        elif i == "Crop":
            title.append('ショート丈')
        elif i == "Knit":
            title.append('ニット')
        elif i == "Dad":
            title.append('ダッド')
        elif i == "Trunks":
            title.append('トランクス')
        elif i == "Swim":
            title.append('水着')
        elif i == "Boardshorts":
            title.append('水着パンツ')
        elif i == "Multipack":
            title.append('マルチパック')
        elif i == "Trunk":
            title.append('トランクス')
        elif i == "Lightweight":
            title.append('軽量')
        elif i == "Down":
            title.append('ダウン')
        elif i == "Sherpa-Lined":
            title.append('裏ボア')
        elif i == "Vest":
            title.append('ベスト')
        elif i == "Mockneck":
            title.append('モックネック')
        elif i == "Colorblock":
            title.append('カラーブロック')
        elif i == "Embroidered ":
            title.append('刺繍入り')
        elif i == "Trucker ":
            title.append('トラッカー')
        elif i == "Ripped":
            title.append('ダメージ')
        elif i == "Straight":
            title.append('ストレート')
        elif i == "Slim ":
            title.append('スリム')
        elif i == "Fleece":
            title.append('フリース')
        elif i == "Super":
            title.append('スーパー')
        elif i == "Crewneck":
            title.append('クルーネック')
        elif i == "Icon":
            title.append('アイコン')
        elif i == "Graphic":
            title.append('グラフィック')
        elif i == "V-Neck":
            title.append('Vネック')
        elif i == "T-Shirt":
            title.append('Tシャツ')
        elif i == "Henley":
            title.append('ヘンリー')
        elif i == "Crop":
            title.append('ショート丈')
        elif i == "Jeans":
            title.append('ジーンズ')
        elif i == "Nylon":
            title.append('ナイロン')
        elif i == "Short-Sleeve":
            title.append('半袖')
        elif i == "Polo":
            title.append('ポロシャツ')
        elif i == "Button-Front":
            title.append('ボタン')
        elif i == "Puffer":
            title.append('ダウン')
        elif i == "Cotton":
            title.append('コットン')
        else:
            title.append(i)


    title = ' '.join(title)
    count = 0
    for c in title:
        if unicodedata.east_asian_width(c) in 'FWA':
            count += 2
        else:
            count += 1
    
    if count > 60:
        ws['A' + str(NewProductNum)].fill = PatternFill(patternType='lightGray')

    return(title)


main()


