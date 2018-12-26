import unittest, time, requests, webbrowser, bs4, datetime, schedule, sys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import csv, os, re, xlrd, openpyxl
from selenium.webdriver.support.ui import WebDriverWait
import time, datetime, bs4, requests, urllib, xlwt, os, re, xlrd, openpyxl, unicodedata
from openpyxl.styles.fills import PatternFill

 

###############
### 基本設定 ###
###############

#エクセル指定
excelName = 'pacsun'
xl_bk = xlrd.open_workbook(excelName + '.xlsx')
xl_sh = xl_bk.sheet_by_name(excelName)
wb = openpyxl.load_workbook('C:\\Users\\tomoa\\Desktop\\' + excelName + '.xlsx') 
ws = wb.active


#サイス取得開始番号の指定
browser = webdriver.Chrome()
browser.implicitly_wait(40)
time.sleep(1)
browser.get('https://www.google.com/')


def main():
    
    #列指定
    yoko = input('新商品追加番号：')
    yoko = int(yoko) - 1
    folderNum = int(yoko) + 1


    #新商品追加
    print('')
    print('新商品追加')
    print('＝＝＝＝＝＝＝')
    print('')
    while True:
        print('')
        print(yoko)
        try:
            #URLへ移動
            browser.implicitly_wait(40)
            time.sleep(1)
            ITEMURL = xl_sh.cell_value(yoko,4)
            if ITEMURL == '-':
                break
            browser.get(ITEMURL)
            browser.implicitly_wait(40)
            time.sleep(10)

            
            #商品名
            name = browser.find_element_by_xpath('//*[@id="pdpMain"]/div[1]/div[1]/h1').text
            ws['A' + str(int(yoko) + 1)].value = name
            japanese(name, yoko, ws)
            ws['B' + str(int(yoko) + 1)].value = japanese(name, yoko, ws)
            print(name)

            #値段
            price = browser.find_element_by_xpath('//*[@id="pdpMain"]/div[1]/div[1]/div/div[2]').text
            price = price.replace('$', '')
            ws['D' + str(int(yoko) + 1)].value = price
            print(price)

                    
            #カラー選択
            colornum = 1
            while True:
                try:
                    browser.find_element_by_xpath('//*[@id="pdpMain"]/div[1]/div[4]/div[2]/ul/li[1]/div[1]/ul/li[' + str(colornum) + ']').click()
                except:
                    break
                time.sleep(5)


                #色
                if colornum == 1:
                    colorcolum = 'H'
                    transcolor = 'I'
                    sizecolum = 'J'
                elif colornum == 2:
                    colorcolum = 'K'
                    transcolor = 'L'
                    sizecolum = 'M'
                elif colornum == 3:
                    colorcolum = 'N'
                    transcolor = 'O'
                    sizecolum = 'P'
                elif colornum == 4:
                    colorcolum = 'Q'
                    transcolor = 'R'
                    sizecolum = 'S'
                elif colornum == 5:
                    colorcolum = 'T'
                    transcolor = 'U'
                    sizecolum = 'V'
                elif colornum == 6:
                    colorcolum = 'W'
                    transcolor = 'X'
                    sizecolum = 'Y'
                elif colornum == 7:
                    colorcolum = 'Z'
                    transcolor = 'AA'
                    sizecolum = 'AB'
                elif colornum == 8:
                    colorcolum = 'AC'
                    transcolor = 'AD'
                    sizecolum = 'AE'
                elif colornum == 9:
                    colorcolum = 'AF'
                    transcolor = 'AG'
                    sizecolum = 'AH'
                elif colornum == 10:
                    colorcolum = 'AI'
                    transcolor = 'AJ'
                    sizecolum = 'AK'


                #色
                color = browser.find_element_by_xpath('//*[@id="pdpMain"]/div[1]/div[4]/div[2]/ul/li[1]/div[2]').text
                ws[colorcolum + str(int(yoko) + 1)].value = color
                colorsTranslate(color)
                ws[transcolor + str(int(yoko) + 1)].value = colorsTranslate(color)
                print(color)

                
                colornum  += 1


                
            #画像
            num = 1
            try:
                while True:
                    if num == 6:
                        break
                    img = browser.find_element_by_xpath('//*[@id="pdpThumbs"]/div/div/div[' + str(num) + ']/div/img')
                    src = img.get_attribute('src')
                    urllib.request.urlretrieve(src, 'C:\\Users\\tomoa\\Desktop\\' + excelName + '\\' + str(folderNum) + '\\' + colorcolum + str(num) + ".jpg")
                    num += 1
                print('画像取得完了')
            except:
                print('画像取得完了')

            
            

                

            #ブランド名
            brandName(name)
            ws['AM' + str(int(yoko) + 1)].value = brandName(name)

    
            #ステータス
            ws['F' + str(int(yoko) + 1)].value = 'new'
            wb.save('C:\\Users\\tomoa\\Desktop\\uo.xlsx')
            yoko += 1
            print('-----------------------')

            
        except IndexError:
            break

        except:
            ws['F' + str(int(yoko) + 1)].value = 'エラー'
            wb.save('C:\\Users\\tomoa\\Desktop\\uo.xlsx')
            yoko += 1
            print('エラー')
            print('-----------------------')


    sys.exit()




########################################################################################################################
########################################################################################################################


# 色の翻訳
def colorsTranslate(color):
    
    if "lack" in color:
        return("ブラック")
    elif "&" in color:
        return("マルチ")
    elif "White" in color:
        return("ホワイト")
    elif "Cream" in color:
        return("ホワイト")
    elif "Brown" in color:
        return("ブラウン")
    elif "Chocolate" in color or 'Chestnut' in color or 'Mustard' in color:
        return("ブラウン")
    elif "Ivory" in color:
        return("グレー")
    elif "Tan" in color:
        return("グレー")
    elif "Navy" in color:
        return("ネイビー")
    elif "Teal" in color:
        return("ネイビー")
    elif "Gold" in color:
        return("イエロー")
    elif "Yellow" in color:
        return("イエロー")
    elif "Green" in color:
        return("グリーン")
    elif "Lime" in color:
        return("グリーン")
    elif "Olive" in color:
        return("グリーン")
    elif "Pink" in color:
        return("ピンク")
    elif "Peach" in color:
        return("ピンク")
    elif "lue" in color:
        return("ブルー")
    elif "Turquoise" in color:
        return("グリーン")
    elif "Denim" in color:
        return("ブルー")
    elif "Sky" in color or 'Slate' in color:
        return("ブルー")
    elif "Grey" in color or 'Lilac' in color:
        return("グレー")
    elif "Charcoal" in color:
        return("グレー")
    elif "Taupe" in color:
        return("グレー")
    elif "Red" in color or 'Rust' in color or 'Crimson' in color:
        return("レッド")
    elif "Maroon" in color:
        return("レッド")
    elif "Coral" in color:
        return("レッド")
    elif "Rose" in color:
        return("レッド")
    elif "Orange" in color:
        return("オレンジ")
    elif "Purple" in color:
        return("パープル")
    elif "Violet" in color:
        return("パープル")
    elif "Lavender" in color:
        return("パープル")
    elif "Beige" in color or 'Khaki' in color:
        return("ベージュ")
    elif "Indigo" in color:
        return("ブルー")
    elif "Assorted" in color:
        return("マルチカラー")
    elif "Mauve" in color:
        return("マルチカラー")
    elif "Multi" in color:
        return("マルチカラー")
    elif "Berry" in color:
        return("ピンク")
    elif "Neutral" in color or 'Birch' in color:
        return("ホワイト")
    elif "Silver" in color or 'Mint' in color or 'Platinum' in color:
        return("シルバー")
    elif "Honey" in color or 'Mint' in color:
        return("マルチカラー")
    else:
        return("#################")



# 商品ブランド名の取得
def brandName(name):
    if "Tommy" in name:
        return "Tommy Hilfiger(トミーヒルフィガー)"
    elif "Vans" in name:
        return  "VANS(バンズ)"
    elif "Champion" in name:
        return "CHAMPION(チャンピオン)"
    elif "adidas" in name:
        return  "adidas(アディダス)"
    elif "The North Face"in name:
        return  "THE NORTH FACE(ザノースフェイス)"
    elif "Alpha Industries"in name:
        return  "Alpha Industries(アルファインダストリー)"
    elif "Levi" in name:
        return  "Levi's(リーバイス)"
    elif "Kappa" in name:
        return "Kappa(カッパ)"
    elif "FILA" in name:
        return  "FILA(フィラ)"
    elif "Ralph" in name:
        return  "LAUREN RALPH LAUREN(ローレンラルフローレン)"
    elif "Puma" in name:
        return "PUMA(プーマ)"
    elif "Patagonia" in name:
        return  "Patagonia(パタゴニア)"
    elif "Stussy" in name:
        return  "STUSSY(ステューシー)"
    elif "Calvin Klein" in name:
        return  "Calvin Klein(カルバンクライン)"
    elif "Teva" in name:
        return  "Teva(テバ)"
    elif "Nike" in name:
        return  "Nike(ナイキ)"
    elif "Asics" in name:
        return  "asics(アシックス)"
    elif "Converse" in name:
        return  "CONVERSE(コンバース)"
    elif "Dr. Martens" in name:
        return  "Dr Martens(ドクターマーチン)"
    elif "Birkenstock" in name:
        return  "BIRKENSTOCK(ビルケンシュトック)"
    elif "Camper" in name:
        return  "CAMPER(カンペール)"
    elif "Reebok" in name:
        return  "Reebok(リーボック)"
    elif "New Balance" in name:
        return  "New Balance(ニューバランス)"
    elif "Skechers" in name:
        return  "SKECHERS(スケッチャーズ)"
    elif "Timberland" in name:
        return  "Timberland(ティンバーランド)"
    elif "Superga" in name:
        return  "SUPERGA(スペルガ)"
    elif "Jeffrey Campbell" in name:
        return  "Jeffrey Campbell(ジェフリーキャンベル)"
    elif "Rocket Dog" in name:
        return  "ROCKET DOG(ロケットドッグ)"
    elif "Columbia" in name:
        return  "Columbia(コロンビア)"
    elif "Dickies" in name:
        return  "Dickies(ディッキーズ)"
    elif "GUESS" in name:
        return  "Guess(ゲス)"
    elif "Herschel Supply Co" in name:
        return  "Herschel Supply(ハーシェルサプライ)"
    elif "Kapten & Son" in name:
        return  "KAPTEN & SON(キャプテン＆サン)"
    elif "Lacoste" in name:
        return  "LACOSTE(ラコステ)"
    elif "Umbro" in name:
        return  "UMBRO(アンブロ)"
    elif "Katin" in name:
        return  "KATIN(ケイティン)"
    elif "UO" in name:
        return  "Urban Outfitters(アーバンアウトフィッターズ)"
    else:
        return "Urban Outfitters(アーバンアウトフィッターズ)"






def japanese(name, NewProductNum, ws):
    title = []

    split = name.split()
    for i in split:

        if i == "Slide":
            title.append('シャワー')
        elif i == 'Denim':
            title.append('デニム')
        elif i == "Flip-Flop":
            title.append('ビーチサンダル')
        elif i == "Hoodie":
            title.append('パーカー')
        elif i == "Jacket" or i == 'Coat' or i == 'Vest' or i == 'Overshirt':
            title.append('ジャケット')
        elif i == "Tee":
            title.append('Tシャツ')
        elif i == "Jersey":
            title.append('ジャージ')
        elif i == "Shirt":
            title.append('シャツ')
        elif i == "Sock":
            title.append('ソックス/靴下')
        elif i == "Boxer":
            title.append('ボクサーパンツ')
        elif i == "Brief":
            title.append('ブリーフ')
        elif i == "Sweatshirt":
            title.append('スウェットシャツ')
        elif i == "Jogger":
            title.append('ジョガー')
        elif i == "Pants":
            title.append('パンツ')
        elif i == "Skinny":
            title.append('スキニー')
        elif i == "Ankle":
            title.append('アンクル')
        elif i == "Shorts":
            title.append('ショートパンツ')
        elif i == "Hat":
            title.append('キャップ')
        elif i == "Flannel":
            title.append('フランネル')
        elif i == "Print":
            title.append('プリント')
        elif i == "Waffle":
            title.append('ワッフル')
        elif i == "Beanie":
            title.append('ニット帽')
        elif i == "Sweatpants":
            title.append('スウェットパンツ')
        elif i == "Gazelle":
            title.append('ガゼル')
        elif i == "Sneaker":
            title.append('スニーカー')
        elif i == "Superstar":
            title.append('スーパースター')
        elif i == "Sandals":
            title.append('サンダル')
        elif i == "Fur":
            title.append('ファー')
        elif i == "Leather":
            title.append('レザー')
        elif i == "Vintage":
            title.append('ビンテージ')
        elif i == "Suede":
            title.append('スエード')
        elif i == "Pastel":
            title.append('パステル')
        elif i == "Mule":
            title.append('ミュール')
        elif i == "Heel":
            title.append('ハイヒール')
        elif i == "Boot":
            title.append('ブーツ')
        elif i == "Sunglasses":
            title.append('サングラス')
        elif i == "Bag":
            title.append('バッグ')
        elif i == "Necklace" or i == "Bracelet":
            title.append('ネックレス')
        elif i == "Backpack":
            title.append('バックパック')
        elif i == "Robe":
            title.append('ローブ')
        elif i == "Watch":
            title.append('時計')
        elif i == "Ring":
            title.append('リング') 
        elif i == "Tank":
            title.append('タンク')
        elif i == "Top":
            title.append('トップ')
        elif i == "Overall":
            title.append('オーバーオール')
        elif i == "Leggings":
            title.append('レギンス')
        elif i == "Track":
            title.append('トラック')
        elif i == "High-Rise":
            title.append('ハイライズ')
        elif i == "Active":
            title.append('アクティブ')
        elif i == "Mini":
            title.append('ミニ')
        elif i == "Twil":
            title.append('ツイル')
        elif i == "Bralette":
            title.append('ブラ')
        elif i == "Hipster":
            title.append('ヒップスターショーツ')
        elif i == "Bra":
            title.append('ブラ')
        elif i == "Undie":
            title.append('下着パンツ')
        elif i == "Cotton":
            title.append('コットン')
        elif i == "Thong":
            title.append('下着パンツ')
        elif i == "Tanga":
            title.append('下着パンツ')
        elif i == "Body":
            title.append('ボディ')
        elif i == "Bikini":
            title.append('ビキニ')
        elif i == "Top":
            title.append('トップ')
        elif i == "Bottom":
            title.append('ボトムパンツ')
        elif i == "One-Piece":
            title.append('ワンピース')
        elif i == "Swimsuit":
            title.append('水着')
        elif i == "Duffle":
            title.append('ダッフル')
        elif i == "Stripe":
            title.append('ストライプ')
        elif i == "Belt":
            title.append('ベルト')
        elif i == "Crossbody":
            title.append('ショルダー')
        elif i == "Tote":
            title.append('トート')
        elif i == "Canvas":
            title.append('キャンバス')
        elif i == "Wedge":
            title.append('ウェッジ')
        elif i == "Button-Down":
            title.append('ボタン')
        elif i == "Satin":
            title.append('サテン')
        elif i == "Fleece":
            title.append('フリース')
        elif i == "Logo":
            title.append('ロゴ')
        elif i == "Pocket":
            title.append('ポケット')
        elif i == "Bomber":
            title.append('ボンバー')
        elif i == "Hooded":
            title.append('フード付き')
        elif i == "Parka":
            title.append('パーカー')
        elif i == "Coach":
            title.append('コーチ')
        elif i == "Windbreaker":
            title.append('ウィンドブレーカー')
        elif i == "Popover":
            title.append('ポップオーバー')
        elif i == "Polo":
            title.append('ポロ')
        elif i == "Boxer":
            title.append('ボクサー')
        elif i == "Brief":
            title.append('ブリーフ')
        elif i == "Floral":
            title.append('花柄')
        elif i == "Full-Zip":
            title.append('ジップアップ')
        elif i == "Set":
            title.append('セット')
        elif i == "Sleep":
            title.append('パジャマ')
        elif i == "Flannele":
            title.append('フランネル')
        elif i == "Plaid":
            title.append('チェック')
        elif i == "Pocket":
            title.append('ポケット')
        elif i == "Checkerboard":
            title.append('チェック')
        elif i == "Patterned":
            title.append('柄入り')
        elif i == "Oxford ":
            title.append('オックスフォード')
        elif i == "Poplin ":
            title.append('アイコン')
        elif i == "Faux":
            title.append('フォックス')
        elif i == "Crop":
            title.append('ショート丈')
        elif i == "Knit":
            title.append('ニット')
        elif i == "Dad":
            title.append('ダッド')
        elif i == "Trunks":
            title.append('トランクス')
        elif i == "Swim":
            title.append('水着')
        elif i == "Boardshorts":
            title.append('水着パンツ')
        elif i == "Multipack":
            title.append('マルチパック')
        elif i == "Trunk":
            title.append('トランクス')
        elif i == "Lightweight":
            title.append('軽量')
        elif i == "Down":
            title.append('ダウン')
        elif i == "Sherpa-Lined":
            title.append('裏ボア')
        elif i == "Vest":
            title.append('ベスト')
        elif i == "Mockneck":
            title.append('モックネック')
        elif i == "Colorblock":
            title.append('カラーブロック')
        elif i == "Embroidered ":
            title.append('刺繍入り')
        elif i == "Trucker ":
            title.append('トラッカー')
        elif i == "Ripped":
            title.append('ダメージ')
        elif i == "Straight":
            title.append('ストレート')
        elif i == "Slim ":
            title.append('スリム')
        elif i == "Fleece":
            title.append('フリース')
        elif i == "Super":
            title.append('スーパー')
        elif i == "Crewneck":
            title.append('クルーネック')
        elif i == "Icon":
            title.append('アイコン')
        elif i == "Graphic":
            title.append('グラフィック')
        elif i == "V-Neck":
            title.append('Vネック')
        elif i == "T-Shirt":
            title.append('Tシャツ')
        elif i == "Henley":
            title.append('ヘンリー')
        elif i == "Crop":
            title.append('ショート丈')
        elif i == "Jeans":
            title.append('ジーンズ')
        elif i == "Nylon":
            title.append('ナイロン')
        elif i == "Short-Sleeve":
            title.append('半袖')
        elif i == "Polo":
            title.append('ポロシャツ')
        elif i == "Button-Front":
            title.append('ボタン')
        elif i == "Puffer":
            title.append('ダウン')
        elif i == "Cotton":
            title.append('コットン')
        else:
            title.append(i)


    title = ' '.join(title)
    count = 0
    for c in title:
        if unicodedata.east_asian_width(c) in 'FWA':
            count += 2
        else:
            count += 1
    
    if count > 60:
        ws['A' + str(NewProductNum + 1)].fill = PatternFill(patternType='lightGray')

    return(title)

main()
