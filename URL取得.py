import unittest, time, requests, webbrowser, bs4, datetime, schedule, sys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import csv, os, re, xlrd, openpyxl
from selenium.webdriver.support.ui import WebDriverWait
import time, datetime, bs4, requests, urllib, xlwt, os, re, xlrd, openpyxl, unicodedata
from openpyxl.styles.fills import PatternFill



def main():
    spreadSheetPath = 'spreadsheets\\'
    #サイス取得開始番号の指定
    browser = webdriver.Chrome()
    LIST = [] 

    #エクセル書き込み設定
    excel = input('excel名:')
    wb = openpyxl.load_workbook(spreadSheetPath + excel + '.xlsx') 
    ws = wb.active

    NewProductNum = input('エクセル追加番号：')


    #URL取得
    if excel == 'backcountry':
        a = 'https://www.backcountry.com/new-arrivals?p=onsaleUS%3Afalse%7Ccategory%3A2.bcs.Men%27s%5C+Clothing.Men%27s%5C+Jackets&nf=1'
        b = 'https://www.backcountry.com/new-arrivals?p=onsaleUS%3Afalse%7Ccategory%3A2.bcs.Women%27s%5C+Clothing.Women%27s%5C+Jackets&nf=1'
        urls = [a,b]
    elif excel == 'bloomingdale':
        a = 'https://www.bloomingdales.com/shop/mens/coats-jackets?id=11548&cm_sp=NAVIGATION-_-TOP_NAV-_-11548-Clothing-Coats-%26-Jackets'
        b = 'https://www.bloomingdales.com/shop/womens-apparel/coats-jackets?id=1001520&cm_sp=NAVIGATION-_-TOP_NAV-_-1001520-Clothing-Coats-%26-Jackets'
        urls = [a,b]
    elif excel == 'anthropologie':
        a = 'https://www.anthropologie.com/dresses'
        b = 'https://www.anthropologie.com/jackets-coats'
        c = 'https://www.anthropologie.com/new-sweaters?feature-product-ids=AN-4114326951628-000,AN-4114089862714-000,AN-49458623-000,AN-4114238540007-000,AN-4114089869967-000,AN-4114339214284-000,AN-4114326959962-000,AN-4113529100970-000'
        d = 'https://www.anthropologie.com/new-sweaters?feature-product-ids=AN-4113529100970-000,AN-4114089862714-000,AN-4114089869967-000,AN-4114238540007-000,AN-4114326951628-000,AN-4114326959962-000,AN-4114339214284-000,AN-49458623-000&page=2'
        e = 'https://www.anthropologie.com/tops-sweatshirts'
        f = 'https://www.anthropologie.com/new-skirts'
        g = 'https://www.anthropologie.com/tops-blouses'
        h = 'https://www.anthropologie.com/tops-tees'
        urls = [a,c,d,e,f,g,h]
    elif excel == 'superdry':
        a = 'https://www.superdry.com/us/mens/jackets'
        b = 'https://www.superdry.com/us/womens/jackets'
        urls = [a,b]
    elif excel == 'pacsun':
        a = 'https://www.pacsun.com/mens/tees/long-sleeve-tees/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CHUF%7CGuess%7CPuma%7CRVCA%7CTommy%20Jeans%7CVans'
        b = 'https://www.pacsun.com/mens/sweaters/'
        c = 'https://www.pacsun.com/mens/jackets-coats/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CConverse%7CKappa%7CLevi%27s%7CRVCA%7CTimberland%7CTommy%20Jeans%7CVans'
        d = 'https://www.pacsun.com/mens/joggers/?prefn1=brand&prefv1=adidas%7CChampion%7CConverse%7CKappa%7CRVCA'
        e = 'https://www.pacsun.com/mens/hoodies/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CConverse%7CGuess%7CHUF%7CKappa%7Ck-Swiss%7CLevi%27s%7CLacoste%7CPuma%7CVans%7CTommy%20Jeans%7CRVCA'
        f = 'https://www.pacsun.com/womens/graphic-t-shirts-tank-tops/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CGuess%7CLacoste%7Ck-Swiss%7CTommy%20Hilfiger%7CTommy%20Jeans'
        g = 'https://www.pacsun.com/womens/hoodies-fleece/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CKappa%7Ck-Swiss%7CPuma%7CTommy%20Hilfiger%7CTommy%20Jeans%7CVans%7CVolcom%7CGuess%7CSuperdry'
        h = 'https://www.pacsun.com/womens/jackets-coats/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CLevi%27s%7CRVCA%7CTommy%20Hilfiger%7CVans%7CTommy%20Jeans%7CKappa'
        i = 'https://www.pacsun.com/womens/shoes/sneakers/?prefn1=brand&prefv1=adidas%7CConverse%7CVans%7CPuma'
        urls = [a,b,c,d,e,f,g,h,i]
 
    for url in urls:
        browser.get(url)
        browser.implicitly_wait(40)
        time.sleep(5)
        
        #エクセルの全URLをリストにする
        NUM = 1
        URLLIST = []
        xl_bk = xlrd.open_workbook(spreadSheetPath + excel + ".xlsx")
        xl_sh = xl_bk.sheet_by_name(excel)
        while True:
            try:
                if not xl_sh.cell_value(NUM,4) in URLLIST:
                    URLLIST.append(xl_sh.cell_value(NUM,4))
                NUM += 1
            except IndexError:
                break


        num = 4
        while True:
            print(NewProductNum)
            try:
                if excel == 'backcountry':
                    name = browser.find_element_by_xpath('/html/body/div[3]/div[2]/div[4]/div[1]/section/div[3]/div[' + str(num) + ']/div[1]/a')
                    href = name.get_attribute('href')
                    print(href)            
                    if not href in LIST and not href in URLLIST:
                        LIST.append(href)
                        ws['E' + str(NewProductNum)].value = href
                        category(url)
                        ws['C' + str(NewProductNum)].value = category(url)
                        wb.save(spreadSheetPath + excel + '.xlsx')
                        NewProductNum = int(NewProductNum) + 1
                elif excel == 'bloomingdale':
                    name = browser.find_element_by_xpath('/html/body/div[3]/div/div/div[1]/div/div/div[2]/div[2]/ul/li/div/ul/li[' + str(num) + ']/div/a')
                    href = name.get_attribute('href')
                    print(href)
                    if not href in LIST and not href in URLLIST:
                        ws['E' + str(NewProductNum)].value = href
                        category(url)
                        ws['C' + str(NewProductNum)].value = category(url)
                        wb.save(spreadSheetPath + excel + '.xlsx')
                        NewProductNum = int(NewProductNum) + 1
                elif excel == 'anthropologie':
                    name = browser.find_element_by_xpath('/html/body/div[1]/div[3]/div[2]/div[2]/div[2]/div[2]/div[2]/div[' + str(num) + ']/span/div[2]/a')
                    href = name.get_attribute('href')
                    href = 'https://www.anthropologie.com' + href
                    print(href)
                    if not href in LIST and not href in URLLIST:
                        ws['E' + str(NewProductNum)].value = href
                        category(url)
                        ws['C' + str(NewProductNum)].value = category(url)
                        wb.save(spreadSheetPath + excel + '.xlsx')
                        NewProductNum = int(NewProductNum) + 1
                elif excel == 'superdry':
                    name = browser.find_element_by_xpath('/html/body/div[4]/div/div[1]/div[4]/div/div[' + str(num) + ']/a')
                    href = name.get_attribute('href')
                    print(href)
                    if not href in LIST and not href in URLLIST:
                        ws['E' + str(NewProductNum)].value = href
                        category(url)
                        ws['C' + str(NewProductNum)].value = category(url)
                        wb.save(spreadSheetPath + excel + '.xlsx')
                        NewProductNum = int(NewProductNum) + 1
                elif excel == 'pacsun':
                    name = browser.find_element_by_xpath('/html/body/div[1]/div[4]/div[6]/div/ul/li[' + str(num) + ']/div/div[2]/div[1]/a')
                    href = name.get_attribute('href')
                    print(href)
                    if not href in LIST and not href in URLLIST:
                        ws['E' + str(NewProductNum)].value = href
                        category(url)
                        ws['C' + str(NewProductNum)].value = category(url)
                        wb.save(spreadSheetPath + excel + '.xlsx')
                        NewProductNum = int(NewProductNum) + 1

                time.sleep(3)
                num = int(num) + 1


            except:
                time.sleep(3)
                break
                
            



def category(url):
    if url == 'https://www.backcountry.com/new-arrivals?p=onsaleUS%3Afalse%7Ccategory%3A2.bcs.Men%27s%5C+Clothing.Men%27s%5C+Jackets&nf=1':
        return('メンズ　ジャケット')
    elif url == 'https://www.backcountry.com/new-arrivals?p=onsaleUS%3Afalse%7Ccategory%3A2.bcs.Women%27s%5C+Clothing.Women%27s%5C+Jackets&nf=1':
        return('レディース　ジャケット')
    elif url == 'https://www.bloomingdales.com/shop/mens/coats-jackets?id=11548&cm_sp=NAVIGATION-_-TOP_NAV-_-11548-Clothing-Coats-%26-Jackets':
        return('メンズ　ジャケット')
    elif url == 'https://www.bloomingdales.com/shop/womens-apparel/coats-jackets?id=1001520&cm_sp=NAVIGATION-_-TOP_NAV-_-1001520-Clothing-Coats-%26-Jackets':
        return('レディース　ジャケット')
    elif url == 'https://www.anthropologie.com/dresses':
        return('レディース　ドレス')
    elif url == 'https://www.anthropologie.com/jackets-coats':
        return('レディース　ジャケット')
    elif url == 'https://www.superdry.com/us/mens/jackets':
        return('メンズ　ジャケット')
    elif url == 'https://www.superdry.com/us/womens/jackets':
        return('レディース　ジャケット')
    elif url == 'https://www.pacsun.com/mens/jackets-coats/':
        return('メンズファッション　アウター・ジャケット　ジャケットその他')
    elif url =='https://www.pacsun.com/womens/jackets-coats/':
        return('レディース　ジャケット')
    elif url =='https://www.anthropologie.com/new-sweaters?feature-product-ids=AN-4114326951628-000,AN-4114089862714-000,AN-49458623-000,AN-4114238540007-000,AN-4114089869967-000,AN-4114339214284-000,AN-4114326959962-000,AN-4113529100970-000':
        return('レディース　セーター')
    elif url =='https://www.anthropologie.com/new-sweaters?feature-product-ids=AN-4113529100970-000,AN-4114089862714-000,AN-4114089869967-000,AN-4114238540007-000,AN-4114326951628-000,AN-4114326959962-000,AN-4114339214284-000,AN-49458623-000&page=2':
        return('レディース　セーター')
    elif url =='https://www.anthropologie.com/tops-sweatshirts':
        return('レディース　スウェット')
    elif url =='https://www.anthropologie.com/new-skirts':
        return('レディース　スカート')
    elif url =='https://www.anthropologie.com/tops-blouses':
        return('レディース　シャツ')
    elif url =='https://www.anthropologie.com/tops-tees':
        return('レディース　T')
    elif url =='https://www.pacsun.com/mens/tees/long-sleeve-tees/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CHUF%7CGuess%7CPuma%7CRVCA%7CTommy%20Jeans%7CVans':
        return('メンズファッション　トップス　Tシャツ・カットソー')
    elif url == 'https://www.pacsun.com/mens/sweaters/':
        return('メンズファッション　トップス　ニット・セーター')
    elif url == 'https://www.pacsun.com/mens/jackets-coats/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CConverse%7CKappa%7CLevi%27s%7CRVCA%7CTimberland%7CTommy%20Jeans%7CVans':
        return('メンズファッション　アウター・ジャケット　ジャケットその他')
    elif url == 'https://www.pacsun.com/mens/joggers/?prefn1=brand&prefv1=adidas%7CChampion%7CConverse%7CKappa%7CRVCA':
        return('メンズファッション　ボトムス　パンツ')
    elif url == 'https://www.pacsun.com/mens/hoodies/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CConverse%7CGuess%7CHUF%7CKappa%7Ck-Swiss%7CLevi%27s%7CLacoste%7CPuma%7CVans%7CTommy%20Jeans%7CRVCA':
        return('メンズファッション　トップス　パーカー・フーディー')
    elif url == 'https://www.pacsun.com/womens/graphic-t-shirts-tank-tops/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CGuess%7CLacoste%7Ck-Swiss%7CTommy%20Hilfiger%7CTommy%20Jeans':
        return('レディース　トップス　Tシャツ・カットソー)
    elif url == 'https://www.pacsun.com/womens/hoodies-fleece/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CKappa%7Ck-Swiss%7CPuma%7CTommy%20Hilfiger%7CTommy%20Jeans%7CVans%7CVolcom%7CGuess%7CSuperdry':
        return('レディース　トップス　パーカー・フーディ')
    elif url == 'https://www.pacsun.com/womens/jackets-coats/?prefn1=brand&prefv1=adidas%7CCalvin%20Klein%7CChampion%7CLevi%27s%7CRVCA%7CTommy%20Hilfiger%7CVans%7CTommy%20Jeans%7CKappa':
        return('レディース　アウター　ジャケット')
    elif url == 'https://www.pacsun.com/womens/shoes/sneakers/?prefn1=brand&prefv1=adidas%7CConverse%7CVans%7CPuma':
        return('レディース　靴・シューズ　スニーカー')
main()


