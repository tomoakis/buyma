import unittest, time, requests, webbrowser, bs4, datetime, schedule
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import csv, os, re, xlrd, openpyxl
from selenium.webdriver.support.ui import WebDriverWait
import time, datetime, bs4, requests, urllib, xlwt, os, re, xlrd, openpyxl, unicodedata
from openpyxl.styles.fills import PatternFill



###############
### 基本設定 ###
###############

#エクセル指定
excelName = 'anthropologie'
xl_bk = xlrd.open_workbook(excelName + '.xlsx')
xl_sh = xl_bk.sheet_by_name(excelName)
wb = openpyxl.load_workbook('C:\\Users\\tomoa\\Desktop\\' + excelName + '.xlsx') 
ws = wb.active


#サイス取得開始番号の指定
browser = webdriver.Firefox()
browser.implicitly_wait(40)
time.sleep(1)
browser.get('https://www.google.com/')


def main():
    
    #列指定
    read_yoko = input('サイズ取得開始番号：')
    read_yoko = int(read_yoko) - 1

    yoko = input('新商品追加番号：')
    yoko = int(yoko) - 1
    folderNum = int(yoko) + 1


    #新商品追加
    print('')
    print('新商品追加')
    print('＝＝＝＝＝＝＝')
    print('')
    while True:
        print('')
        print(yoko + 1)
        try:
        #URLへ移動
            browser.implicitly_wait(40)
            time.sleep(1)
            ITEMURL = xl_sh.cell_value(yoko,4)
            browser.get(ITEMURL)
            browser.implicitly_wait(40)
            time.sleep(10)

            
            #商品名
            name = browser.find_element_by_xpath('//*[@id="u-skip-anchor"]/div[2]/div[2]/div[1]/div[2]/div[2]/div/div/div[2]/div[1]/h1/span').text
            ws['A' + str(int(yoko) + 1)].value = name
            japanese(name, yoko, ws)
            ws['B' + str(int(yoko) + 1)].value = 'アンソロ★' + japanese(name, yoko, ws)
            print(name)

            #値段
            price = browser.find_element_by_xpath('//*[@id="u-skip-anchor"]/div[2]/div[2]/div[1]/div[2]/div[2]/div/div/div[2]/div[1]/div/p[1]/span/span').text
            price = price.replace('$', '')
            ws['D' + str(int(yoko) + 1)].value = price
            print(price)

            #色
            color = browser.find_element_by_xpath('//*[@id="product-add-form"]/div[1]/div[1]/div[1]/div/fieldset/legend/span').text
            ws['H' + str(int(yoko) + 1)].value = color
            colorsTranslate(color)
            ws['I' + str(int(yoko) + 1)].value = colorsTranslate(color)
            print(color)

            
    ##        #サイズ
    ##        sizebox = []
    ##        sizeList = []
    ##        size = browser.find_element_by_xpath('//*[@id="product-add-form"]/div[1]/div[1]/div[2]/div[2]/div/div/fieldset/ul').text
    ##        
    ##        sizebox.append(size)
    ##        if sizebox == ['XS\nS\nM\nL\nXL\nXXL']:
    ##            stop = 7
    ##        elif sizebox == ['XXS\nXS\nS\nM\nL\nXL']:
    ##            stop = 7
    ##        elif sizebox == ['XS\nS\nM\nL\nXL']:
    ##            stop = 6
    ##        elif sizebox == ['XS\nS\nM\nL']:
    ##            stop = 5
    ##        elif sizebox == ['S\nM\nL\nXL']:
    ##            stop = 5
    ##        elif sizebox == ['ONE SIZE']:
    ##            stop = 2
    ##        elif sizebox == ['26W\n28W\n29W\n30W\n31W\n32W\n33W\n34W\n36W\n38W']:
    ##            stop = 11
    ##        elif sizebox == ['26W\n28W\n29W\n30W\n31W\n32W\n33W\n34W\n36W']:
    ##            stop = 10
    ##        elif sizebox == ['28W\n30W\n31W\n32W\n33W\n34W\n36W\n38W']:
    ##            stop = 9
    ##        elif sizebox == ['00\n0\n1\n3\n5\n7\n9\n11\n13\n15']:
    ##            stop = 11
    ##        elif sizebox == ['5/6\n7/8\n9/10\n11/12\n13/14\n15/16']:
    ##            stop = 7
    ##        elif sizebox == ['00\n0\n2\n4\n6\n8\n10\n12\n14\n16']:
    ##            stop = 11
    ##
    ##            
    ##        num = 1
    ##        while True:
    ##            if num == stop:
    ##                break
    ##            size = browser.find_element_by_xpath('//*[@id="product-add-form"]/div[1]/div[1]/div[2]/div/div/div/fieldset/ul/li[' + str(num) + ']')
    ##            sizes = size.get_attribute('class')
    ##            if 'c-radio-styled__small' in sizes and not 'is-disabled' in sizes:
    ##                SIZE = browser.find_element_by_xpath('//*[@id="product-add-form"]/div[1]/div[1]/div[2]/div/div/div/fieldset/ul/li[' + str(num) + ']/label').text
    ##                sizeList.append(SIZE)
    ##
    ##            num += 1
    ##                
    ##
    ##        sizeList = ','.join(sizeList)
    ##        ws['J' + str(int(yoko) + 1)].value = sizeList
    ##        print(sizeList)
            
            #画像
            num = 1
            while True:
                if num == 5:
                    break
                try:
                    img = browser.find_element_by_xpath('//*[@id="slider-thumbnail__slide-inner-' + str(num) + '"]/img')
                    src = img.get_attribute('src')
                    src = src.replace('150&qlt=80&fit=constrain', '900&qlt=80&fit=constrain')
                    urllib.request.urlretrieve(src, 'C:\\Users\\tomoa\\Desktop\\anthropologie\\' + str(folderNum) + '\\' + str(num) + ".jpg")
                    num += 1
                except:
                    break
            print('画像取得完了')
            


            #ステータス
            ws['F' + str(int(yoko) + 1)].value = 'new'
            wb.save('C:\\Users\\tomoa\\Desktop\\anthropologie.xlsx')
            yoko += 1
            folderNum += 1
            print('-----------------------')

                
        except IndexError:
            break

        except:
            ws['F' + str(int(yoko) + 1)].value = 'new'
            ws['AM' + str(int(yoko) + 1)].value = 'anthropologie'
            wb.save('C:\\Users\\tomoa\\Desktop\\anthropologie.xlsx')
            yoko += 1
            folderNum = yoko + 1
            print('エラー')
            print('-----------------------')







########################################################################################################################
########################################################################################################################


# 色の翻訳
def colorsTranslate(color):
    
    if "LACK" in color:
        return("ブラック")
    elif "&" in color:
        return("マルチ")
    elif "WHITE" in color:
        return("ホワイト")
    elif "CREAM" in color:
        return("ホワイト")
    elif "BROWN" in color:
        return("ブラウン")
    elif "CHOCOLATE" in color or 'CHESTNUT' in color or 'MUSTERD' in color:
        return("ブラウン")
    elif "IVORY" in color:
        return("グレー")
    elif "TAN" in color:
        return("グレー")
    elif "NAVY" in color:
        return("ネイビー")
    elif "TEAL" in color:
        return("ネイビー")
    elif "GOLD" in color:
        return("イエロー")
    elif "YELLOW" in color:
        return("イエロー")
    elif "GREEN" in color:
        return("グリーン")
    elif "LIME" in color:
        return("グリーン")
    elif "OLIVE" in color:
        return("グリーン")
    elif "PINK" in color:
        return("ピンク")
    elif "PEACH" in color:
        return("ピンク")
    elif "BLUE" in color:
        return("ブルー")
    elif "TURQANTHROPOLOGIESE" in color:
        return("グリーン")
    elif "DENIM" in color:
        return("ブルー")
    elif "SKY" in color or 'SLATE' in color:
        return("ブルー")
    elif "GREY" in color or 'LILAC' in color:
        return("グレー")
    elif "CHARCOAL" in color:
        return("グレー")
    elif "TAUPE" in color:
        return("グレー")
    elif "RED" in color or 'RUST' in color or 'CRIMSON' in color:
        return("レッド")
    elif "MAROON" in color:
        return("レッド")
    elif "CORAL" in color:
        return("レッド")
    elif "ROSE" in color:
        return("レッド")
    elif "ORANGE" in color:
        return("オレンジ")
    elif "PURPLE" in color:
        return("パープル")
    elif "VIOLET" in color:
        return("パープル")
    elif "LAVENDER" in color:
        return("パープル")
    elif "BEIGE" in color or 'KHAKI' in color:
        return("ベージュ")
    elif "INDIGO" in color:
        return("ブルー")
    elif "ASSORTED" in color:
        return("マルチカラー")
    elif "MAUVE" in color:
        return("マルチカラー")
    elif "MULTI" in color:
        return("マルチカラー")
    elif "BERRY" in color:
        return("ピンク")
    elif "NEUTRAL" in color or 'BIRCH' in color:
        return("ホワイト")
    elif "SILVER" in color or 'MINT' in color or 'PLATINUM' in color:
        return("シルバー")
    elif "HONEY" in color or 'MINT' in color:
        return("マルチカラー")
    else:
        return("#################")


def japanese(name, NewProductNum, ws):
    title = []

    split = name.split()
    for i in split:

        if i == "Slide":
            title.append('シャワー')
        elif i == 'Denim':
            title.append('デニム')
        elif i == "Flip-Flop":
            title.append('ビーチサンダル')
        elif i == "Hoodie":
            title.append('パーカー')
        elif i == "Jacket" or i == 'Coat' or i == 'Vest' or i == 'Overshirt':
            title.append('ジャケット')
        elif i == "Tee":
            title.append('Tシャツ')
        elif i == "Jersey":
            title.append('ジャージ')
        elif i == "Shirt":
            title.append('シャツ')
        elif i == "Sock":
            title.append('ソックス/靴下')
        elif i == "Boxer":
            title.append('ボクサーパンツ')
        elif i == "Brief":
            title.append('ブリーフ')
        elif i == "Sweatshirt":
            title.append('スウェットシャツ')
        elif i == "Jogger":
            title.append('ジョガー')
        elif i == "Pants":
            title.append('パンツ')
        elif i == "Skinny":
            title.append('スキニー')
        elif i == "Ankle":
            title.append('アンクル')
        elif i == "Shorts":
            title.append('ショートパンツ')
        elif i == "Hat":
            title.append('キャップ')
        elif i == "Flannel":
            title.append('フランネル')
        elif i == "Print":
            title.append('プリント')
        elif i == "Waffle":
            title.append('ワッフル')
        elif i == "Beanie":
            title.append('ニット帽')
        elif i == "Sweatpants":
            title.append('スウェットパンツ')
        elif i == "Gazelle":
            title.append('ガゼル')
        elif i == "Sneaker":
            title.append('スニーカー')
        elif i == "Superstar":
            title.append('スーパースター')
        elif i == "Sandals":
            title.append('サンダル')
        elif i == "Fur":
            title.append('ファー')
        elif i == "Leather":
            title.append('レザー')
        elif i == "Vintage":
            title.append('ビンテージ')
        elif i == "Suede":
            title.append('スエード')
        elif i == "Pastel":
            title.append('パステル')
        elif i == "Mule":
            title.append('ミュール')
        elif i == "Heel":
            title.append('ハイヒール')
        elif i == "Boot":
            title.append('ブーツ')
        elif i == "Sunglasses":
            title.append('サングラス')
        elif i == "Bag":
            title.append('バッグ')
        elif i == "Necklace" or i == "Bracelet":
            title.append('ネックレス')
        elif i == "Backpack":
            title.append('バックパック')
        elif i == "Robe":
            title.append('ローブ')
        elif i == "Watch":
            title.append('時計')
        elif i == "Ring":
            title.append('リング') 
        elif i == "Tank":
            title.append('タンク')
        elif i == "Top":
            title.append('トップ')
        elif i == "Overall":
            title.append('オーバーオール')
        elif i == "Leggings":
            title.append('レギンス')
        elif i == "Track":
            title.append('トラック')
        elif i == "High-Rise":
            title.append('ハイライズ')
        elif i == "Active":
            title.append('アクティブ')
        elif i == "Mini":
            title.append('ミニ')
        elif i == "Twil":
            title.append('ツイル')
        elif i == "Bralette":
            title.append('ブラ')
        elif i == "Hipster":
            title.append('ヒップスターショーツ')
        elif i == "Bra":
            title.append('ブラ')
        elif i == "Undie":
            title.append('下着パンツ')
        elif i == "Cotton":
            title.append('コットン')
        elif i == "Thong":
            title.append('下着パンツ')
        elif i == "Tanga":
            title.append('下着パンツ')
        elif i == "Body":
            title.append('ボディ')
        elif i == "Bikini":
            title.append('ビキニ')
        elif i == "Top":
            title.append('トップ')
        elif i == "Bottom":
            title.append('ボトムパンツ')
        elif i == "One-Piece":
            title.append('ワンピース')
        elif i == "Swimsuit":
            title.append('水着')
        elif i == "Duffle":
            title.append('ダッフル')
        elif i == "Stripe":
            title.append('ストライプ')
        elif i == "Belt":
            title.append('ベルト')
        elif i == "Crossbody":
            title.append('ショルダー')
        elif i == "Tote":
            title.append('トート')
        elif i == "Canvas":
            title.append('キャンバス')
        elif i == "Wedge":
            title.append('ウェッジ')
        elif i == "Button-Down":
            title.append('ボタン')
        elif i == "Satin":
            title.append('サテン')
        elif i == "Fleece":
            title.append('フリース')
        elif i == "Logo":
            title.append('ロゴ')
        elif i == "Pocket":
            title.append('ポケット')
        elif i == "Bomber":
            title.append('ボンバー')
        elif i == "Hooded":
            title.append('フード付き')
        elif i == "Parka":
            title.append('パーカー')
        elif i == "Coach":
            title.append('コーチ')
        elif i == "Windbreaker":
            title.append('ウィンドブレーカー')
        elif i == "Popover":
            title.append('ポップオーバー')
        elif i == "Polo":
            title.append('ポロ')
        elif i == "Boxer":
            title.append('ボクサー')
        elif i == "Brief":
            title.append('ブリーフ')
        elif i == "Floral":
            title.append('花柄')
        elif i == "Full-Zip":
            title.append('ジップアップ')
        elif i == "Set":
            title.append('セット')
        elif i == "Sleep":
            title.append('パジャマ')
        elif i == "Flannele":
            title.append('フランネル')
        elif i == "Plaid":
            title.append('チェック')
        elif i == "Pocket":
            title.append('ポケット')
        elif i == "Checkerboard":
            title.append('チェック')
        elif i == "Patterned":
            title.append('柄入り')
        elif i == "Oxford ":
            title.append('オックスフォード')
        elif i == "Poplin ":
            title.append('アイコン')
        elif i == "Faux":
            title.append('フォックス')
        elif i == "Crop":
            title.append('ショート丈')
        elif i == "Knit":
            title.append('ニット')
        elif i == "Dad":
            title.append('ダッド')
        elif i == "Trunks":
            title.append('トランクス')
        elif i == "Swim":
            title.append('水着')
        elif i == "Boardshorts":
            title.append('水着パンツ')
        elif i == "Multipack":
            title.append('マルチパック')
        elif i == "Trunk":
            title.append('トランクス')
        elif i == "Lightweight":
            title.append('軽量')
        elif i == "Down":
            title.append('ダウン')
        elif i == "Sherpa-Lined":
            title.append('裏ボア')
        elif i == "Vest":
            title.append('ベスト')
        elif i == "Mockneck":
            title.append('モックネック')
        elif i == "Colorblock":
            title.append('カラーブロック')
        elif i == "Embroidered ":
            title.append('刺繍入り')
        elif i == "Trucker ":
            title.append('トラッカー')
        elif i == "Ripped":
            title.append('ダメージ')
        elif i == "Straight":
            title.append('ストレート')
        elif i == "Slim ":
            title.append('スリム')
        elif i == "Fleece":
            title.append('フリース')
        elif i == "Super":
            title.append('スーパー')
        elif i == "Crewneck":
            title.append('クルーネック')
        elif i == "Icon":
            title.append('アイコン')
        elif i == "Graphic":
            title.append('グラフィック')
        elif i == "V-Neck":
            title.append('Vネック')
        elif i == "T-Shirt":
            title.append('Tシャツ')
        elif i == "Henley":
            title.append('ヘンリー')
        elif i == "Crop":
            title.append('ショート丈')
        elif i == "Jeans":
            title.append('ジーンズ')
        elif i == "Nylon":
            title.append('ナイロン')
        elif i == "Short-Sleeve":
            title.append('半袖')
        elif i == "Polo":
            title.append('ポロシャツ')
        elif i == "Button-Front":
            title.append('ボタン')
        elif i == "Puffer":
            title.append('ダウン')
        elif i == "Cotton":
            title.append('コットン')
        else:
            title.append(i)


    title = ' '.join(title)
    count = 0
    for c in title:
        if unicodedata.east_asian_width(c) in 'FWA':
            count += 2
        else:
            count += 1
    
    if count > 60:
        ws['A' + str(NewProductNum + 1)].fill = PatternFill(patternType='lightGray')

    return(title)


main()
