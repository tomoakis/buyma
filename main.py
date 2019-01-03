# -*- coding: utf-8 -*-
import unittest
import time
import requests
import webbrowser
import datetime
import os
import re
import xlrd
import openpyxl
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.alert import Alert

spreadSheetPath = './spreadsheets/'
photoPath = ''


### エクセル指定 ###
###################
# name = 'エクセル名：'
excelName = input('xlsx name: ')
xl_bk = xlrd.open_workbook(spreadSheetPath + excelName + '.xlsx')
xl_sh = xl_bk.sheet_by_name(excelName)
wb = openpyxl.load_workbook(spreadSheetPath + excelName + '.xlsx')
ws = wb.active


### 出品開始番号指定 ###
######################
# name = '出品開始番号：'
yoko = input('Row number: ')
yoko = int(yoko) - 1
folderNum = int(yoko) + 1


###　ログイン　###
#################
browser = webdriver.Chrome(executable_path='./drivers/chromedriver')
browser.get("https://www.buyma.com/my/sell/new/")
email = browser.find_element_by_id('txtLoginId')
email.send_keys('namitaketomi123@gmail.com')
password = browser.find_element_by_id('txtLoginPass')
password.send_keys('seasider0093')
browser.find_element_by_id('login_do').click()
browser.implicitly_wait(40)
time.sleep(1)
browser.get("https://www.buyma.com/my/sell/new/")

while True:
    print(folderNum)
    try:

        ### J列読み込み ###
        ##################
        action = xl_sh.cell_value(yoko, 5)

        ### 新商品出品 ###
        #################
        if action == 'new':

            # ###ブランド名###
            # ###############
            brand = xl_sh.cell_value(yoko, 38)
            # brandName = browser.find_element_by_xpath('/html/body/div[3]/div[2]/div[1]/div/div/div/div/div/div[2]/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[1]/div/div/div/div/input')
            # brandName.send_keys('アバクロ')
            # browser.implicitly_wait(40)
            # time.sleep(1)

            ###商品名###
            ############
            productName = xl_sh.cell_value(yoko,1)
            browser.find_element_by_xpath('/html/body/div[3]/div[2]/div[1]/div/div/div/div/div/div[2]/form/div[2]/div[1]/div/div[2]/div/div/div[1]/input').send_keys(productName)
            time.sleep(1)


            ###　商品写真　###
            #################
            dir = excelName + '\\' + str(folderNum)
            files = os.listdir(dir) #ファイルのリストを取得

            for file in files:
                picture = os.path.abspath(excelName + '\\' + str(folderNum) + '\\' + file)
                images = browser.find_element_by_xpath('/html/body/div[3]/div[2]/div[1]/div/div/div/div/div/div[2]/form/div[1]/div/div/div[2]/div/div/div[1]/div/div/div/input')
                browser.implicitly_wait(40)
                time.sleep(1)
                browser.execute_script("arguments[0].style.display = 'block';", images)
                browser.implicitly_wait(40)
                time.sleep(1)
                images.send_keys(picture)
                browser.implicitly_wait(40)
                time.sleep(1)



            ### 商品コメント ###
            ###################
            productName = xl_sh.cell_value(yoko,0)
            
            if 'ホリスター' in brand:
                comment = '大人気のホリスターから「' + productName + '」をお届けします。ホリスター・カンパニーとは、アメリカのカジュアルファッションブランドです。2000年にアバクロンビーアンドフィッチ社によって設立されたブランドであり、世界中に６００以上の店舗を構えるトップブランドの１つです。\n\n「カモメ」をモチーフにしたブランドロゴが特徴的で、アメリカ西海岸のサーファースタイルをイメージしたデザインとなっています。アメリカでは姉妹ブランドの「アバクロンビーアンドフィッチ」「ルール No ９２５」や「アメリカンイーグル」等と並んで、若い世代に絶大な人気を誇るブランドです。\n\n略してホリスターと呼ばれることが多く、アメリカの調査によると１０代の若者に２番目に人気のあるファッションブランドであり、世界的にも絶大な人気を博しているカジュアルブランドです。\n\n基本、注文後の買い付けです。\n\n在庫に限りがあり、店舗の出品回転も速いためオンライン・店舗完売の時がよくあります。\n\n●サイズなどについては、商品が手元にない場合そのため正確な数字をお知らせできないことがあります。公式サイトに記載されているサイズをそのまま記載しておりますので、そちらを参考にして頂けると幸いです。\n\n●注文後早ければ翌日、最大1週間ほどお時間かかることもあります。\n（店舗にて売れ切れの場合はオンラインで発注します）\n\n●発送方法は、基本アメリカからファーストクラス便で発送します。\n発送後、到着までに早ければ１週間、税関や空輸が混雑していますと２週間-３週間掛かることもあります。\n\n●直接店舗で買い付けた場合は商品に、店舗で使われている香水の匂い、多少のヨレ感がありますこと予めご了承ください。\n\n●商品発送前に入念に検品をして発送することを徹底して心がけております。\n\n●商品の在庫数が極限られていますので、受注時に既に売れ切れている場合がございます。その場合にはキャンセルという形で対応させていただきますのでご理解ください。\n（バイマよりご返金）'
            elif 'アバクロ' in brand:
                comment = '大人気のアバクロから「' + productName + '」をお届けします。\n\n基本、注文後の買い付けです。\n\n在庫に限りがあり、店舗の出品回転も速いためオンライン・店舗完売の時がよくあります。\n\n●サイズなどについては、商品が手元にない場合そのため正確な数字をお知らせできないことがあります。公式サイトに記載されているサイズをそのまま記載しておりますので、そちらを参考にして頂けると幸いです。\n\n●注文後早ければ翌日、最大1週間ほどお時間かかることもあります。\n（店舗にて売れ切れの場合はオンラインで発注します）\n\n●発送方法は、基本アメリカからファーストクラス便で発送します。\n発送後、到着までに早ければ１週間、税関や空輸が混雑していますと２週間-３週間掛かることもあります。\n\n●直接店舗で買い付けた場合は商品に、店舗で使われている香水の匂い、多少のヨレ感がありますこと予めご了承ください。\n\n●商品発送前に入念に検品をして発送することを徹底して心がけております。\n\n●商品の在庫数が極限られていますので、受注時に既に売れ切れている場合がございます。その場合にはキャンセルという形で対応させていただきますのでご理解ください。\n（バイマよりご返金）'
            else:
                comment = '大人気の「' + productName + '」をお届けします。\n\n（店舗にて売れ切れの場合はオンラインで発注します）\n●発送方法は、基本アメリカからファーストクラス便で発送します。\n発送後、到着までに早ければ１週間、税関や空輸が混雑していますと２週間-３週間掛かることもあります。\n●直接店舗で買い付けた場合は商品に、店舗で使われている香水の匂い、多少のヨレ感がありますこと予めご了承ください。\n●商品発送前に入念に検品をして発送することを徹底して心がけております。\n●商品の在庫数が極限られていますので、受注時に既に売れ切れている場合がございます。その場合にはキャンセルという形で対応させていただきますのでご理解ください。\n（バイマよりご返金)'

            browser.find_element_by_xpath('/html/body/div[3]/div[2]/div[1]/div/div/div/div/div/div[2]/form/div[2]/div[2]/div/div[2]/div/div/div[1]/textarea').send_keys(comment)
            time.sleep(1)





            #####　カテゴリ　####
            ####################
            category = xl_sh.cell_value(yoko,2)

            # 第1段落
            browser.find_element_by_xpath('//*[@id="react-select-2--value"]').click()
            browser.implicitly_wait(40)
            time.sleep(1)
            CATEGORY1 = browser.find_element_by_xpath('//*[@id="react-select-2--value"]/div[2]')
            browser.execute_script("arguments[0].style.display = 'block';", CATEGORY1)
            time.sleep(1)


# CATEGORY1 = browser.find_element_by_class_name('Select-menu-outer')
# browser.execute_script("arguments[0].style.display = 'block';", CATEGORY1)
# time.sleep(1)
            # if 'メンズ' in category:
            #     CATEGORY1.send_keys('レディースファッション')
            # elif 'レディースファッション' in category:
            #     CATEGORY1.send_keys('レディースファッション')
            # elif 'キッズ' in category:
            #     CATEGORY1.send_keys('キッズ')
            # browser.implicitly_wait(40)
            # time.sleep(1)
            
            # #第2段落
            # CATEGORY2 = browser.find_element_by_xpath('/html/body/div[3]/div[2]/div[1]/div/div/div/div/div/div[2]/form/div[3]/div[1]/div/div[2]/div/div/div[1]/div/div[2]/div/div/span[1]/div[2]')
            # if 'アウター' in category:
            #     CATEGORY2.send_keys('アウター・ジャケット')
            # browser.implicitly_wait(40)
            # time.sleep(1)
            
            # #第3段落
            # CATEGORY3 = browser.find_element_by_xpath('/html/body/div[3]/div[2]/div[1]/div/div/div/div/div/div[2]/form/div[3]/div[1]/div/div[2]/div/div/div[1]/div/div[3]/div/div/span[1]/div[2]')
            # if 'ジャケット' in category:
            #     CATEGORY2.send_keys('ジャケットその他')
            # browser.implicitly_wait(40)
            # time.sleep(1)




            ### 色とサイズ ###
            #################
            white = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[1]/span[1]'
            black = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[2]/span[1]'
            grey = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[3]/span[1]'
            brown = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[4]/span[1]'
            beju = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[5]/span[1]'
            green = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[6]/span[1]'
            blue = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[7]/span[1]'
            navy = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[8]/span[1]'
            purple = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[9]/span[1]'
            yellow = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[10]/span[1]'
            pink = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[11]/span[1]'
            red = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[12]/span[1]'
            orange = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[13]/span[1]'
            silver = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[14]/span[1]'
            gold = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[15]/span[1]'
            clear = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[16]/span[1]'
            multiColor = '//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[1]/div/a[17]/span[1]'


            # 設定するボタンをクリック
            browser.find_element_by_xpath('/html/body/div[3]/div[2]/div[1]/div/div[1]/div/form/div[2]/div[3]/table/tbody/tr[6]/td/div[1]/div[1]/a[1]').click()
            browser.implicitly_wait(40)

            # 複数色を選択、入力
            color = [8,11,14,17,20,23,26,29,32,35]
            for i in color:
                color = xl_sh.cell_value(yoko,i)
                colorName = xl_sh.cell_value(yoko,i - 1)
                if color == '':
                    break


                # 色アイコンの指定
                if color == 'ホワイト':
                    browser.find_element_by_xpath(white).click()
                elif color == 'ブラック':
                    browser.find_element_by_xpath(black).click()
                elif color == 'グレー':
                    browser.find_element_by_xpath(grey).click()
                elif color == 'ブラウン':
                    browser.find_element_by_xpath(brown).click()
                elif color == 'ベージュ':
                    browser.find_element_by_xpath(beju).click()
                elif color == 'グリーン':
                    browser.find_element_by_xpath(green).click()
                elif color == 'ブルー':
                    browser.find_element_by_xpath(blue).click()
                elif color == 'ネイビー':
                    browser.find_element_by_xpath(navy).click()
                elif color == 'パープル':
                    browser.find_element_by_xpath(purple).click()
                elif color == 'イエロー':
                    browser.find_element_by_xpath(yellow).click()
                elif color == 'ピンク':
                    browser.find_element_by_xpath(pink).click()
                elif color == 'レッド':
                    browser.find_element_by_xpath(red).click()
                elif color == 'オレンジ':
                    browser.find_element_by_xpath(orange).click()
                elif color == 'シルバー':
                    browser.find_element_by_xpath(silver).click()
                elif color == 'ゴールド':
                    browser.find_element_by_xpath(gold).click()
                elif color == 'クリア':
                    browser.find_element_by_xpath(clear).click()
                elif color == 'マルチカラー':
                    browser.find_element_by_xpath(multiColor).click()
                    
                # 色アイコンの名称を記入
                browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[2]/div/input').send_keys(colorName)
                browser.implicitly_wait(40)
                time.sleep(2)
                
                # 追加するボタンをクリック
                browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[3]/div/div/div/div[2]/div/a').click()




            # 全色の重複しない全サイズをリストにまとめる
            newSizeLIST = []
            newSizecolum = [9,12,15,18,21,24,27,30,33,36]
            for i in newSizecolum:
                newSize = xl_sh.cell_value(yoko,i)
                if newSize == '':
                    lennewSize = len(newSizeLIST)
                    break

                # サイズをリストとしてスプリットさせる
                newSize = newSize.split(',')
                # もしリストにサイズがなければ追加
                for i in newSize:

                    # sortでサイズ順に並ぶように一時変換
                    if i == 'XS':
                        i = 'AS'
                    elif i == 'XXS':
                        i = 'AAS'
                    elif i == 'S':
                        i = 'B'
                    elif i == 'M':
                        i = 'C'
                    elif i == 'L':
                        i = 'D'
                    elif i == '5/6':
                        i = '1'
                    elif i == '7/8':
                        i = '11'
                    elif i == '9/10':
                        i = '11/'

                    if not i in newSizeLIST:
                        newSizeLIST.append(i)


            # 一番多いサイズ数分ボックスを追加する
            count = 1
            while True:
                if count == lennewSize:
                    break
                browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/div[4]/a').click()
                browser.implicitly_wait(40)
                time.sleep(1)
                count += 1


            # サイズを反映
            sizecheckBox = []
            sizeNumber = 2
            newSizeLIST.sort()
            for i in newSizeLIST:
                # サイズを元に変換
                if i == 'AS':
                    i = 'XS'
                elif i == 'AAS':
                    i = 'XXS'
                elif i == 'B':
                    i = 'S'
                elif i == 'C':
                    i = 'M'
                elif i == 'D':
                    i = 'L'
                elif i == '1':
                    i = '5/6'
                elif i == '11':
                    i = '7/8'
                elif i == '11/':
                    i = '9/10'

                if not i in sizecheckBox:
                    sizecheckBox.append(i)


                # サイズを反映
                browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[2]/input').send_keys(i)
                browser.implicitly_wait(40)
                time.sleep(1)

                # 日本参考サイズ(洋服)
                if category == 'メンズファッション 靴・ブーツ・サンダル 靴・ブーツ・サンダルその他' or category == 'メンズファッション 靴・ブーツ・サンダル スニーカー' or category == 'メンズファッション 靴・ブーツ・サンダル サンダル':
                    if i == '4' or i == '4.5' or i == '5':
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '5.5' or i == '35 1/2':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("23.5cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '6' or i == '36':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("24cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '6.5' or i == '36 1/2':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("24.5cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '7' or i == '37' or i == 'M 7/W 8.5' or i == 'US 7/EU 40':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("25cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '7.5' or i == '37 1/2' or i == 'M 7.5/W 9':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("25.5cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '8' or i == '38' or i == 'M 8/W 9.5' or i == 'US 8/EU 41':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("26cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '8.5' or i == '38 1/2' or i == 'M 8.5/W 10':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("26.5cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '9' or i == '39' or i == 'M 9/W 10.5' or i == 'US 9/EU 42':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("27cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '9.5' or i == '39 1/2' or i == 'M 9.5/W 11' or i == 'US 9.5/EU 43':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("27.5cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '10' or i == '40' or i == 'M 10/W 11.5' or i == 'US 10/EU 44':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("28cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '10.5' or i == '40 1/2' or i == 'M 10.5/W 12':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("28.5cm")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == '11' or i == '41' or i == 'M 11/W 12.5' or i == 'US 11/EU 45':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("29cm以上")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == 'M 11.5/W 13':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("29cm以上")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == 'M 12/W 13.5' or i == 'US 12/EU 46' or i == '12':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("29cm以上")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == 'M 12.5/W 14' or i == '12.5':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("29cm以上")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == 'M 13/W 14.5' or i == '13':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("29cm以上")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == 'M 13.5/W 15' or i == '13.5':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("29cm以上")
                        browser.implicitly_wait(40)
                        time.sleep(1)
                    elif i == 'M 14/W 15.5' or i == '14':
                        browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("29cm以上")
                        browser.implicitly_wait(40)
                        time.sleep(1)



                elif i == 'XXS' or i == '2' or i == '24':
                    browser.implicitly_wait(40)
                    time.sleep(1)
                elif i == 'XS' or i == '4' or i == '25' or i == 'XS/S' or i == '30/32':
                    browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("S")
                    browser.implicitly_wait(40)
                    time.sleep(1)
                elif i == 'S' or i == '6' or i == '26' or i == 'S/M' or i == '32/32':
                    browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("M")
                    browser.implicitly_wait(40)
                    time.sleep(1)
                elif i == 'M' or i == '8' or i == '27' or i == 'M/L' or i == '34/32':
                    browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("L")
                    browser.implicitly_wait(40)
                    time.sleep(1)
                elif i == 'L' or i == 'XL' or i == 'XXL' or i == 'L/XL' or i == '10' or i == '12' or i == '14' or i == '28' or i == '29' or i == '30' or i == '31' or i == '32' or i == '36/32':
                    browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("XL以上")
                    browser.implicitly_wait(40)
                    time.sleep(1)
                elif i == 'ONE SIZE':
                    browser.find_element_by_xpath(' //*[@id="rdoSelectSize2"]').click()
                    browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys("ONE SIZE")
                    break
                else:
                    browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(sizeNumber) + ']/td[3]/select').send_keys(i)
                    browser.implicitly_wait(40)
                    time.sleep(1)

                sizeNumber += 1



            # 全色の色のチェックを外す
            for i in newSizecolum:
                newSize = xl_sh.cell_value(yoko,i)
                if newSize == '':
                    break

                newSize = newSize.split(',')

                if i == 9:
                    yokocheck = '4'
                elif i == 12:
                    yokocheck = '5'
                elif i == 15:
                    yokocheck = '6'
                elif i == 18:
                    yokocheck = '7'
                elif i == 21:
                    yokocheck = '8'
                elif i == 24:
                    yokocheck = '9'
                elif i == 27:
                    yokocheck = '10'
                elif i == 30:
                    yokocheck = '11'
                elif i == 33:
                    yokocheck = '12'
                elif i == 36:
                    yokocheck = '13'

                tatecheck = 2
                for i in sizecheckBox:
                    browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(tatecheck) + ']/td[' + yokocheck + ']/input').click()
                    tatecheck += 1


            # 色ごとにサイズボックスの色を入れる
            for i in newSizecolum:
                newSize = xl_sh.cell_value(yoko,i)
                if newSize == '':
                    break
                newSize = newSize.split(',')

                if i == 9:
                    yokocheck = '4'
                elif i == 12:
                    yokocheck = '5'
                elif i == 15:
                    yokocheck = '6'
                elif i == 18:
                    yokocheck = '7'
                elif i == 21:
                    yokocheck = '8'
                elif i == 24:
                    yokocheck = '9'
                elif i == 27:
                    yokocheck = '10'
                elif i == 30:
                    yokocheck = '11'
                elif i == 33:
                    yokocheck = '12'
                elif i == 36:
                    yokocheck = '13'



                for ii in newSize:
                    tatecheck = 2

                    for a in sizecheckBox:
                        if a == ii:
                            browser.find_element_by_xpath('//*[@id="components"]/div/div[1]/div[4]/table/tbody/tr[' + str(tatecheck) + ']/td[' + yokocheck + ']/input').click()
                        tatecheck += 1





            # 全サイズ反映後、設定ボタンをクリック
            browser.find_element_by_xpath('//*[@id="components"]/div/div[2]/a[2]').click()



            #サイズ補足説明欄###
            ##################
            category = xl_sh.cell_value(yoko,2)
            if brand == 'Hollister Co.(ホリスター)':
                if 'メンズ' in category:
                    sizeComment = "★メンズ参考サイズ★\n\n基本注文確認後、買い付けしています。手元に在庫がありませんので下記の参考サイズを参考にして頂ければ幸いです。\n\nサイズに不安な場合は、注文後、買い付けが完了しだい実寸の平置きをお知らせすることはできます。確認後のサイズ変更も可能です。\n\n（胸囲cm）\nS　91 - 96\nM　97 - 101\nL　102 - 106\nXL 107 - 111\n\n（袖長さcm）\nS　82 - 85\nM　86 - 87\nL　89 - 90\nXL 91 - 93\n\n\n（ウェスト）\nXS 28 (71cm)\nS　 29 - 30 (74-76cm)\nM　 31 - 32 (79-81cm)\nL　 33 - 34 (84-86cm)\nXL 36 (89cm)\n\n(足のサイズ）\nS　26.5ｃｍ\nM　27.5ｃｍ\nL　28.3ｃｍ\nXL　29.1ｃｍ\n\nその他アバクロ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816/\nその他ホリスター商品はこちら→https://www.buyma.com/r/_HOLLISTER-%E3%83%9B%E3%83%AA%E3%82%B9%E3%82%BF%E3%83%BC/-B4256816/\nその他アバクロキッズ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816F1/%E5%A4%A7%E4%BA%BA%E3%82%82OK/"
                elif 'レディース' in category:
                    sizeComment = "★レディース参考サイズ★\n\n基本注文確認後、買い付けしています。手元に在庫がありませんので下記の参考サイズを参考にして頂ければ幸いです。\n\nサイズに不安な場合は、注文後、買い付けが完了しだい実寸の平置きをお知らせすることはできます。確認後のサイズ変更も可能です。\n\n（胸囲cm）\nXS　80 - 84　（5 - 7号）\nS　86 - 89　　（９号）\nM　91 - 94　　（１１号）\nL　96 - 97　　（１３号）\n\n\n（ウェスト INCHES）\nXS　23 - 25　（5 - 7号）\nS　26 - 27　 （7 - 9号）\nM　28 - 29　 （9 - 11号）\nL　30 - 31　 （11 - 13号）\n\n(足のサイズ）\nXS　23.2ｃｍ\nS　23.8ｃｍ\nM　24.8ｃｍ\nL　25.4ｃｍ\nXL　29.1ｃｍ\n\nその他アバクロ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816/\nその他ホリスター商品はこちら→https://www.buyma.com/r/_HOLLISTER-%E3%83%9B%E3%83%AA%E3%82%B9%E3%82%BF%E3%83%BC/-B4256816/\nその他アバクロキッズ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816F1/%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD%E3%82%AD%E3%83%83%E3%82%BA/"
            elif brand == 'Abercrombie & Fitch(アバクロ)':
                if 'メンズ' in category:
                    sizeComment = "★メンズ参考サイズ★\n\n基本注文確認後、買い付けしています。手元に在庫がありませんので下記の参考サイズを参考にして頂ければ幸いです。\n\nサイズに不安な場合は、注文後、買い付けが完了しだい実寸の平置きをお知らせすることはできます。確認後のサイズ変更も可能です。\n\n（胸囲cm）\nS　91 - 96\nM　97 - 101\nL　102 - 106\nXL 107 - 111\n\n（袖長さcm）\nS　82 - 85\nM　86 - 87\nL　89 - 90\nXL 91 - 93\n\n\n（ウェスト）\nXS 28 (71cm)\nS　 29 - 30 (74-76cm)\nM　 31 - 32 (79-81cm)\nL　 33 - 34 (84-86cm)\nXL 36 (89cm)\n\n(足のサイズ）\nS　26.5ｃｍ\nM　27.5ｃｍ\nL　28.3ｃｍ\nXL　29.1ｃｍ\n\nその他アバクロ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816/\nその他ホリスター商品はこちら→https://www.buyma.com/r/_HOLLISTER-%E3%83%9B%E3%83%AA%E3%82%B9%E3%82%BF%E3%83%BC/-B4256816/\nその他アバクロキッズ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816F1/%E5%A4%A7%E4%BA%BA%E3%82%82OK/"
                elif 'レディース' in category:
                    sizeComment = "★レディース参考サイズ★\n\n基本注文確認後、買い付けしています。手元に在庫がありませんので下記の参考サイズを参考にして頂ければ幸いです。\n\nサイズに不安な場合は、注文後、買い付けが完了しだい実寸の平置きをお知らせすることはできます。確認後のサイズ変更も可能です。\n\n（胸囲cm）\nXS　80 - 84　（5 - 7号）\nS　86 - 89　　（９号）\nM　91 - 94　　（１１号）\nL　96 - 97　　（１３号）\n\n\n（ウェスト INCHES）\nXS　23 - 25　（5 - 7号）\nS　26 - 27　 （7 - 9号）\nM　28 - 29　 （9 - 11号）\nL　30 - 31　 （11 - 13号）\n\n(足のサイズ）\nXS　23.2ｃｍ\nS　23.8ｃｍ\nM　24.8ｃｍ\nL　25.4ｃｍ\nXL　29.1ｃｍ\n\nその他アバクロ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816/\nその他ホリスター商品はこちら→https://www.buyma.com/r/_HOLLISTER-%E3%83%9B%E3%83%AA%E3%82%B9%E3%82%BF%E3%83%BC/-B4256816/\nその他アバクロキッズ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816F1/%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD%E3%82%AD%E3%83%83%E3%82%BA/"
                elif 'ベビー・キッズ' in category:
                    sizeComment = "★アバクロキッズ参考サイズ★\n\n基本注文確認後、買い付けしています。手元に在庫がありませんので下記の参考サイズを参考にして頂ければ幸いです。\n\n（身長cm）\n5/6　110 - 122\n7/8　122 - 135\n9/10　135 - 145\n11/12　145 - 152\n13/14　152 - 160\n15/16　160 - 165\n\n（胸囲ｃｍ）\n5/6　58 - 64\n7/8　64 - 69\n9/10　69 - 72\n11/12　72 - 76\n13/14　76 - 80\n15/16　80 - 84\n\n(足のサイズ）\n12/13　18.4ｃｍ\n1/2　20.3ｃｍ\n3/4　21.9ｃｍ\n5/6　23.5ｃｍ\n7/8　24.8ｃｍ\n\n※在庫の変動が激しいので、購入前に在庫確認をよろしくお願いします。\n\nその他アバクロ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816/\nその他ホリスター商品はこちら→https://www.buyma.com/r/_HOLLISTER-%E3%83%9B%E3%83%AA%E3%82%B9%E3%82%BF%E3%83%BC/-B4256816/\nその他アバクロキッズ商品はこちら→https://www.buyma.com/r/_%E3%82%A2%E3%83%90%E3%82%AF%E3%83%AD-ABERCROMBIE&FITCH/-B4256816F1/%E5%A4%A7%E4%BA%BA%E3%82%82OK/"
            else:
                    sizeComment = "※在庫の変動が激しいので在庫確認をよろしくお願いします。"
                    browser.find_element_by_xpath('//*[@id="item_color_size"]').send_keys(sizeComment)
                    browser.implicitly_wait(40)
                    time.sleep(1)


            browser.find_element_by_xpath('//*[@id="item_color_size"]').send_keys(sizeComment)
            browser.implicitly_wait(40)
            time.sleep(1)



            ###シーズン###
            #############
            season = "2018-19AW"
            browser.find_element_by_xpath('//*[@id="season"]').send_keys(season)
            browser.implicitly_wait(40)
            time.sleep(1)




##            ### タグ　###
##            ############
# browser.find_element_by_xpath('//*[@id="chkForm"]/div[2]/div[2]/table/tbody/tr[11]/td/div[2]/a').click()
# browser.implicitly_wait(40)
# time.sleep(1)
# browser.find_element_by_xpath('//*[@id="_r_tag_check_419_37"]').click()
# browser.implicitly_wait(40)
# time.sleep(1)
# browser.find_element_by_xpath('//*[@id="r_tag_select_box"]/div/div[6]/ul/li/button').click()
# browser.implicitly_wait(40)
# time.sleep(1)
##
##
##            ###テーマ###
##            ###########
# thema = '円高還元セール特集！'
# browser.find_element_by_xpath('//*[@id="chkForm"]/div[2]/div[2]/table/tbody/tr[12]/td/a').click()
# browser.implicitly_wait(40)
# time.sleep(3)
# browser.find_element_by_xpath('//*[@id="my"]/div[13]/div[2]/div/div[1]/table/tbody/tr[5]').click()
# browser.implicitly_wait(40)
# time.sleep(1)



            ###値段###
            ##########
            if 'シャツ' in category or '帽子' in category  or '水着' in category or 'キッズ用トップス' in category  or '子供用帽子' in category:
                sellingPrice = (((float(xl_sh.cell_value(yoko,3)))*100*1.0685*1.2)+3200)*1.08
                sellingPrice = round(sellingPrice, -2)
                browser.find_element_by_xpath('//*[@id="price"]').send_keys(int(sellingPrice))
                browser.implicitly_wait(40)
                time.sleep(1)
            elif 'デニム' in category:
                sellingPrice = (((float(xl_sh.cell_value(yoko,3)))*100*1.0685*1.2)+4500)*1.08
                sellingPrice = round(sellingPrice, -2)
                browser.find_element_by_xpath('//*[@id="price"]').send_keys(int(sellingPrice))
                browser.implicitly_wait(40)
                time.sleep(1)
            else:
                sellingPrice = (((float(xl_sh.cell_value(yoko,3)))*100*1.0685*1.2)+4000)*1.08
                sellingPrice = round(sellingPrice, -2)
                browser.find_element_by_xpath('//*[@id="price"]').send_keys(int(sellingPrice))
                browser.implicitly_wait(40)
                time.sleep(1)




            ###配送方法###
            #############
            shipping = "179915"
            tracking = '229273'
            sippingId = '//*[@id="shipping-method-checkbox' + str(shipping) + '"]'
            browser.find_element_by_xpath(sippingId).click()
            browser.implicitly_wait(40)
            time.sleep(1)
            trackingID = '//*[@id="shipping-method-checkbox' + str(tracking) + '"]'
            browser.find_element_by_xpath(trackingID).click()
            browser.implicitly_wait(40)
            time.sleep(1)

            '''
            #参考価格###
            ###########
            regularPrice = "190000"
            browser.find_element_by_xpath('//*[@id="itemedit[reference_price_kbn]2"]').click()
            browser.find_element_by_xpath('//*[@id="reference_price"]').send_keys(regularPrice)
            browser.implicitly_wait(40)
            time.sleep(1)
            '''

            ###数量###
            ##########
            productNum = "2"
            browser.find_element_by_xpath('//*[@id="pieces"]').send_keys(productNum)
            browser.implicitly_wait(40)
            time.sleep(1)



            ###買い付け発送地###
            ##################
            browser.find_element_by_xpath('//*[@id="rdoMyActArea2"]').click()
            browser.implicitly_wait(40)
            time.sleep(1)
            browser.find_element_by_xpath('//*[@id="itemedit_purchase_area"]').send_keys('北米')
            browser.implicitly_wait(40)
            time.sleep(1)
            browser.find_element_by_xpath('//*[@id="rdoMyHassoArea2"]').click()
            browser.implicitly_wait(40)
            time.sleep(1)
            browser.find_element_by_xpath('//*[@id="hasso_foreign"]').send_keys('北米')
            browser.implicitly_wait(40)
            time.sleep(1)


            '''
            #ショップ名###
            #############
            shopName = xl_sh.cell_value(yoko,6)
            browser.find_element_by_xpath('//*[@id="itemedit_konyuchi"]').send_keys(shopName)
            browser.implicitly_wait(40)
            time.sleep(1)
            '''

            ###入力内容を確認ボタン###
            ########################
            browser.find_element_by_xpath('//*[@id="confirmButton"]').click()
            browser.implicitly_wait(40)
            time.sleep(1)


            ###出品完了###
            #############
            browser.find_element_by_xpath('//*[@id="done"]').click()
            browser.implicitly_wait(40)
            time.sleep(1)


            ###出品URLを保存###
            ##################
            browser.find_element_by_link_text('出品リストへ戻る').click()
            browser.implicitly_wait(40)
            time.sleep(1)
            ItemURL = browser.find_element_by_xpath('//*[@id="inputform"]/table/tbody/tr[2]/td[4]/p[2]/a[1]').text
            ItemURL = 'https://www.buyma.com/my/sell/new/?iid=' + ItemURL
            browser.implicitly_wait(40)
            time.sleep(1)
            ws['AN' + str(int(yoko) + 1)].value = ItemURL
            wb.save('C:\\Users\\tomoa\\Workspace\\buyma' + excelName + '.xlsx')


            ###出品画面に移動###
            ###################

            browser.get('https://www.buyma.com/my/sell/new/')
            browser.implicitly_wait(40)
            time.sleep(1)


            ###次の行に行く###
            #################
            yoko += 1
            folderNum = int(yoko) + 1




        ### 変更なし ###
        ###############
        else:
            yoko += 1
            folderNum = int(yoko) + 1
            continue




    except IndexError:
        break


    except:
        try:
            ws['F' + str(int(yoko) + 1)].value = 'ERROR'
            wb.save('C:\\Users\\tomoa\\Workspace\\buyma' + excelName + '.xlsx')
            browser.get('https://www.buyma.com/my/sell/new/')
            browser.implicitly_wait(40)
            time.sleep(1)
            yoko += 1
            folderNum = yoko + 1
        except:
            time.sleep(1)
            Alert(browser).accept()
            ws['F' + str(int(yoko) + 1)].value = 'ERROR'
            wb.save('C:\\Users\\tomoa\\Workspace\\buyma' + excelName + '.xlsx')
            browser.get('https://www.buyma.com/my/sell/new/')
            browser.implicitly_wait(40)
            time.sleep(1)
            yoko += 1
            folderNum = yoko + 1
